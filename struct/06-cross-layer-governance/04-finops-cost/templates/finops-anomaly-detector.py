#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FinOps 成本异常检测器 (FinOps Cost Anomaly Detector)

功能：
    1. 读取 YAML / JSON / CSV 格式的历史成本数据（按资源/服务/团队逐日或逐月）。
    2. 实现两种异常检测算法：
       - 基于均值 + 3σ 的 Z-Score
       - 基于环比增长率阈值（如 >30%）
    3. 输出异常列表：资源、服务、团队、日期、实际成本、预期成本、
       偏差百分比、异常类型，并可导出为 CSV / Excel。

命令行用法：
    python finops-anomaly-detector.py --input example-anomaly.yaml --output report.csv
    python finops-anomaly-detector.py --input costs.json --output report.xlsx --method zscore --threshold 2.5
    python finops-anomaly-detector.py --input costs.csv --output report.csv --method growth --threshold 0.30

对齐来源：
    - FinOps Foundation: Cost Anomaly Detection Capability
    - FOCUS (FinOps Open Cost and Usage Specification) 1.0
      https://focus.finops.org/

依赖（均优雅降级）：
    - PyYAML：用于读取 YAML 输入；缺失时仍支持 JSON / CSV。
    - openpyxl：用于 Excel 输出；缺失时自动降级为 CSV。
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

# ---------------------------------------------------------------------------
# 可选依赖适配
# ---------------------------------------------------------------------------
try:
    import yaml  # type: ignore

    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
DEFAULT_ZSCORE_THRESHOLD = 3.0
DEFAULT_GROWTH_THRESHOLD = 0.30
MIN_SAMPLES_ZSCORE = 3
REQUIRED_FIELDS = {"date", "resource", "service", "team", "cost"}


# ---------------------------------------------------------------------------
# 数据模型
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class CostRecord:
    """单条历史成本记录。"""

    date: str
    resource: str
    service: str
    team: str
    cost: float


@dataclass
class Anomaly:
    """检测出的异常记录。"""

    resource: str
    service: str
    team: str
    date: str
    actual_cost: float
    expected_cost: float
    deviation_percent: float
    anomaly_type: str
    zscore: Optional[float] = None
    growth_rate: Optional[float] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "resource": self.resource,
            "service": self.service,
            "team": self.team,
            "date": self.date,
            "actual_cost": self.actual_cost,
            "expected_cost": self.expected_cost,
            "deviation_percent": self.deviation_percent,
            "anomaly_type": self.anomaly_type,
            "zscore": self.zscore,
            "growth_rate": self.growth_rate,
        }


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------
def _records_from_mapping(data: Any) -> List[CostRecord]:
    """从 dict/list 结构解析记录。"""
    if isinstance(data, dict):
        records = data.get("records", [])
    elif isinstance(data, list):
        records = data
    else:
        raise ValueError("输入数据必须是对象（含 records 字段）或记录列表。")

    if not isinstance(records, list):
        raise ValueError("records 字段必须是列表。")

    result: List[CostRecord] = []
    for idx, item in enumerate(records, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"第 {idx} 条记录不是对象。")
        missing = REQUIRED_FIELDS - set(item.keys())
        if missing:
            raise ValueError(f"第 {idx} 条记录缺少字段：{', '.join(sorted(missing))}")
        result.append(
            CostRecord(
                date=str(item["date"]).strip(),
                resource=str(item["resource"]).strip(),
                service=str(item["service"]).strip(),
                team=str(item["team"]).strip(),
                cost=float(item["cost"]),
            )
        )
    return result


def _records_from_csv(path: Path) -> List[CostRecord]:
    """从 CSV 文件解析记录。"""
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError("CSV 文件为空或缺少表头。")
        fieldnames = [h.strip().lower() for h in reader.fieldnames]
        missing = REQUIRED_FIELDS - set(fieldnames)
        if missing:
            raise ValueError(f"CSV 缺少必需列：{', '.join(sorted(missing))}")
        result: List[CostRecord] = []
        for row in reader:
            result.append(
                CostRecord(
                    date=row["date"].strip(),
                    resource=row["resource"].strip(),
                    service=row["service"].strip(),
                    team=row["team"].strip(),
                    cost=float(row["cost"]),
                )
            )
    return result


def load_data(path: Path) -> List[CostRecord]:
    """根据文件后缀自动选择解析器。"""
    suffix = path.suffix.lower()
    if suffix in {".yaml", ".yml"}:
        if not HAS_YAML:
            raise RuntimeError(
                "读取 YAML 需要 PyYAML。请安装：pip install pyyaml"
            )
        with path.open("r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        return _records_from_mapping(data)
    if suffix == ".json":
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return _records_from_mapping(data)
    if suffix == ".csv":
        return _records_from_csv(path)
    raise ValueError(
        f"不支持的输入格式：{suffix}。请使用 .yaml / .yml / .json / .csv。"
    )


# ---------------------------------------------------------------------------
# 异常检测算法
# ---------------------------------------------------------------------------
def _group_records(records: List[CostRecord]) -> Dict[Tuple[str, str, str], List[CostRecord]]:
    """按 (resource, service, team) 分组，并按日期排序。"""
    groups: Dict[Tuple[str, str, str], List[CostRecord]] = defaultdict(list)
    for rec in records:
        groups[(rec.resource, rec.service, rec.team)].append(rec)
    for key in groups:
        groups[key].sort(key=lambda r: r.date)
    return groups


def _mean(values: Sequence[float]) -> float:
    return sum(values) / len(values)


def _stdev(values: Sequence[float]) -> float:
    if len(values) < 2:
        return 0.0
    m = _mean(values)
    variance = sum((x - m) ** 2 for x in values) / (len(values) - 1)
    return math.sqrt(variance)


def detect_zscore(
    records: List[CostRecord],
    threshold: float = DEFAULT_ZSCORE_THRESHOLD,
    min_samples: int = MIN_SAMPLES_ZSCORE,
) -> List[Anomaly]:
    """
    基于历史均值的 Z-Score 异常检测。

    对每个时间点，使用其之前所有历史数据计算均值与标准差，
    若 |Z-Score| > threshold 则标记为异常。
    """
    anomalies: List[Anomaly] = []
    groups = _group_records(records)

    for (resource, service, team), series in groups.items():
        for i in range(min_samples, len(series)):
            history = [r.cost for r in series[:i]]
            current = series[i]
            mean_cost = _mean(history)
            std_cost = _stdev(history)

            if std_cost == 0:
                continue

            zscore = (current.cost - mean_cost) / std_cost
            if abs(zscore) > threshold:
                deviation_percent = (
                    ((current.cost - mean_cost) / mean_cost) * 100
                    if mean_cost != 0
                    else float("inf")
                )
                anomalies.append(
                    Anomaly(
                        resource=resource,
                        service=service,
                        team=team,
                        date=current.date,
                        actual_cost=current.cost,
                        expected_cost=round(mean_cost, 4),
                        deviation_percent=round(deviation_percent, 2),
                        anomaly_type="zscore",
                        zscore=round(zscore, 4),
                    )
                )

    anomalies.sort(key=lambda a: (a.date, a.resource, a.service))
    return anomalies


def detect_growth(
    records: List[CostRecord],
    threshold: float = DEFAULT_GROWTH_THRESHOLD,
) -> List[Anomaly]:
    """
    基于环比增长率阈值异常检测。

    若当期成本相对上期增长超过 threshold（如 0.30 即 30%），
    则标记为异常。
    """
    anomalies: List[Anomaly] = []
    groups = _group_records(records)

    for (resource, service, team), series in groups.items():
        for i in range(1, len(series)):
            current = series[i]
            previous = series[i - 1]
            if previous.cost == 0:
                continue

            growth_rate = (current.cost - previous.cost) / previous.cost
            if growth_rate > threshold:
                deviation_percent = growth_rate * 100
                anomalies.append(
                    Anomaly(
                        resource=resource,
                        service=service,
                        team=team,
                        date=current.date,
                        actual_cost=current.cost,
                        expected_cost=round(previous.cost, 4),
                        deviation_percent=round(deviation_percent, 2),
                        anomaly_type="growth_rate",
                        growth_rate=round(growth_rate, 4),
                    )
                )

    anomalies.sort(key=lambda a: (a.date, a.resource, a.service))
    return anomalies


def detect_anomalies(
    records: List[CostRecord],
    method: str,
    zscore_threshold: float,
    growth_threshold: float,
) -> List[Anomaly]:
    """根据指定方法执行异常检测。"""
    anomalies: List[Anomaly] = []
    if method in ("zscore", "both"):
        anomalies.extend(detect_zscore(records, threshold=zscore_threshold))
    if method in ("growth", "both"):
        anomalies.extend(detect_growth(records, threshold=growth_threshold))
    return anomalies


# ---------------------------------------------------------------------------
# 输出
# ---------------------------------------------------------------------------
def print_anomalies(anomalies: List[Anomaly]) -> None:
    """在控制台打印异常列表。"""
    if not anomalies:
        print("未检测到成本异常。")
        return

    headers = [
        "Resource",
        "Service",
        "Team",
        "Date",
        "Actual",
        "Expected",
        "Deviation%",
        "Type",
    ]
    rows = [
        [
            a.resource,
            a.service,
            a.team,
            a.date,
            f"{a.actual_cost:.2f}",
            f"{a.expected_cost:.2f}",
            f"{a.deviation_percent:.2f}%",
            a.anomaly_type,
        ]
        for a in anomalies
    ]

    widths = [max(len(str(r[i])) for r in [headers] + rows) for i in range(len(headers))]

    def fmt(cells: List[str]) -> str:
        return " | ".join(str(c).ljust(widths[i]) for i, c in enumerate(cells))

    print(fmt(headers))
    print("-" * (sum(widths) + 3 * (len(headers) - 1)))
    for row in rows:
        print(fmt(row))


def write_csv(anomalies: List[Anomaly], path: Path) -> None:
    """导出异常列表为 CSV。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "resource",
        "service",
        "team",
        "date",
        "actual_cost",
        "expected_cost",
        "deviation_percent",
        "anomaly_type",
        "zscore",
        "growth_rate",
    ]
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for anomaly in anomalies:
            writer.writerow(anomaly.to_dict())


def write_excel(anomalies: List[Anomaly], path: Path) -> None:
    """导出异常列表为 Excel（需要 openpyxl）。"""
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "导出 Excel 需要 openpyxl。请安装：pip install openpyxl"
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cost Anomalies"

    headers = [
        "Resource",
        "Service",
        "Team",
        "Date",
        "Actual Cost",
        "Expected Cost",
        "Deviation %",
        "Anomaly Type",
        "Z-Score",
        "Growth Rate",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for anomaly in anomalies:
        ws.append(
            [
                anomaly.resource,
                anomaly.service,
                anomaly.team,
                anomaly.date,
                anomaly.actual_cost,
                anomaly.expected_cost,
                anomaly.deviation_percent,
                anomaly.anomaly_type,
                anomaly.zscore if anomaly.zscore is not None else "",
                anomaly.growth_rate if anomaly.growth_rate is not None else "",
            ]
        )

    for column_cells in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell.border = thin_border
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(path)


def resolve_output_paths(output_arg: Path) -> Tuple[Path, Optional[Path]]:
    """
    根据 --output 参数决定 CSV 与可选 Excel 路径。

    规则：
      - 若输出以 .xlsx 结尾，则同时生成 .xlsx 与同名 .csv；
      - 否则统一按 CSV 处理（自动补全 .csv）。
    """
    suffix = output_arg.suffix.lower()
    if suffix == ".xlsx":
        csv_path = output_arg.with_suffix(".csv")
        return csv_path, output_arg
    if suffix != ".csv":
        output_arg = output_arg.with_suffix(".csv")
    return output_arg, None


# ---------------------------------------------------------------------------
# 命令行入口
# ---------------------------------------------------------------------------
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="FinOps 成本异常检测器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input",
        "-i",
        type=Path,
        required=True,
        help="输入历史成本数据文件（.yaml / .json / .csv）。",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=Path("reports/finops-anomaly-report.csv"),
        help="输出报告路径，默认 reports/finops-anomaly-report.csv。",
    )
    parser.add_argument(
        "--method",
        choices=["zscore", "growth", "both"],
        default="both",
        help="异常检测方法：zscore（均值+3σ）、growth（环比增长率）、both（默认）。",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=None,
        help=(
            "阈值。对 zscore 方法表示标准差倍数（默认 3.0）；"
            "对 growth 方法表示增长率（默认 0.30，即 30%）。"
        ),
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if not args.input.exists():
        print(f"错误：输入文件不存在：{args.input}", file=sys.stderr)
        return 1

    # 阈值语义按方法区分
    if args.method == "zscore":
        zscore_threshold = args.threshold if args.threshold is not None else DEFAULT_ZSCORE_THRESHOLD
        growth_threshold = DEFAULT_GROWTH_THRESHOLD
    elif args.method == "growth":
        zscore_threshold = DEFAULT_ZSCORE_THRESHOLD
        growth_threshold = args.threshold if args.threshold is not None else DEFAULT_GROWTH_THRESHOLD
    else:  # both
        zscore_threshold = args.threshold if args.threshold is not None else DEFAULT_ZSCORE_THRESHOLD
        growth_threshold = args.threshold if args.threshold is not None else DEFAULT_GROWTH_THRESHOLD

    try:
        records = load_data(args.input)
    except Exception as exc:
        print(f"读取输入失败：{exc}", file=sys.stderr)
        return 1

    if not records:
        print("输入文件中没有成本记录。", file=sys.stderr)
        return 1

    anomalies = detect_anomalies(
        records,
        method=args.method,
        zscore_threshold=zscore_threshold,
        growth_threshold=growth_threshold,
    )

    print(f"已加载 {len(records)} 条成本记录，检测到 {len(anomalies)} 条异常。\n")
    print_anomalies(anomalies)

    csv_path, excel_path = resolve_output_paths(args.output)
    try:
        write_csv(anomalies, csv_path)
        print(f"\nCSV 报告已保存：{csv_path}")
    except Exception as exc:
        print(f"\n保存 CSV 失败：{exc}", file=sys.stderr)
        return 1

    if excel_path:
        if HAS_OPENPYXL:
            try:
                write_excel(anomalies, excel_path)
                print(f"Excel 报告已保存：{excel_path}")
            except Exception as exc:
                print(f"保存 Excel 失败：{exc}", file=sys.stderr)
                return 1
        else:
            print(
                f"\n警告：openpyxl 未安装，跳过 Excel 输出。"
                f"仅生成 CSV：{csv_path}",
                file=sys.stderr,
            )

    return 0


if __name__ == "__main__":
    sys.exit(main())
