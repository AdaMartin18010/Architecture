#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FinOps Excel 公式模板生成器 (L1–L8 Cost Allocation Template Generator)

功能：
    1. 使用 openpyxl 生成可复用的 FinOps 成本分摊 Excel 模板
    2. 包含 L1–L8 八个工作表，并使用 Excel 公式实现自动计算：
       - L1-原始成本：云服务原始账单（含 ResourceID、ServiceName、Cost、Tag:Team 等）
       - L2-分摊规则：定义 Team 的分摊键（AllocationKey%）与规则说明
       - L3-分摊结果：使用 SUMIFS + XLOOKUP/VLOOKUP 将 L1 成本按 L2 规则分摊到各 Team
       - L4-单位经济学：每用户/每交易/每 Token 成本计算（含公式）
       - L5-预算预测：12 个月实际成本与预算，含 MoM 增长率、线性预测、移动平均及年度汇总
       - L6-承诺折扣优化：对比按需、RI、SP、Spot 年度成本，推荐最优方案并计算节省
       - L7-异常检测：基于均值/标准差/Z-Score 标记异常成本记录
       - L8-AI GPU 分摊：按团队/项目/模型分摊 GPU 与 Token 成本，并计算占比
    3. 命令行支持 --output 参数，默认输出到 dist/finops-cost-allocation-template.xlsx

用法：
    python finops-excel-template.py
    python finops-excel-template.py --output ./my-template.xlsx

依赖：
    - openpyxl

对齐来源：
    - FinOps Foundation Framework 2026
    - FOCUS (FinOps Open Cost and Usage Specification)
    - 本项目 unit-economics.md / ai-cost-allocation.md 模板
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import List, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# 常量与样例数据
# ---------------------------------------------------------------------------
DEFAULT_OUTPUT = Path("dist/finops-cost-allocation-template.xlsx")

# L1 原始成本样例：模拟一份云服务账单
L1_HEADERS = ["ResourceID", "ServiceName", "Cost", "Tag:Team", "Tag:Env", "Tag:Project"]
L1_SAMPLE_DATA: List[Tuple[str, str, float, str, str, str]] = [
    ("i-0a1b2c3d4e5f", "EC2", 5000.00, "Platform", "prod", "SharedPlatform"),
    ("rds-appa-001", "RDS", 2400.00, "AppA", "prod", "AppA-DB"),
    ("rds-appb-001", "RDS", 1800.00, "AppB", "prod", "AppB-DB"),
    ("eks-control", "EKS", 800.00, "Platform", "prod", "SharedPlatform"),
    ("openai-inference", "AzureOpenAI", 2400.00, "AppA", "prod", "AppA-AI"),
    ("s3-data-lake", "S3", 600.00, "Data", "prod", "DataLake"),
    ("datadog-sub", "SaaS", 1200.00, "Platform", "prod", "Observability"),
    ("lambda-appb", "Lambda", 400.00, "AppB", "prod", "AppB-API"),
    ("cloudfront-cdn", "CloudFront", 900.00, "AppA", "prod", "AppA-Frontend"),
    ("vpc-flow-logs", "VPC", 300.00, "Platform", "prod", "SharedPlatform"),
]

# L2 分摊规则：共享成本池按 Team 的比例进行二次分摊
# 注：L1 中已带 Tag:Team 的直接归属成本会按 Tag 汇总，未标签/共享部分按 L2 比例拆分。
L2_HEADERS = ["Team", "AllocationKey%", "RuleDescription"]
L2_SAMPLE_RULES: List[Tuple[str, float, str]] = [
    ("AppA", 0.35, "按活跃应用数与收入贡献分摊共享平台成本"),
    ("AppB", 0.25, "按活跃应用数分摊共享平台成本"),
    ("Platform", 0.25, "平台团队承担基础设施与可观测性共享成本"),
    ("Data", 0.15, "数据团队承担数据湖与存储共享成本"),
]

# L4 单位经济学样例数据
L4_HEADERS = [
    "Metric",
    "TotalCost",
    "UnitVolume",
    "CostPerUnit",
    "Notes",
]
L4_SAMPLE_DATA: List[Tuple[str, float, float, str]] = [
    ("CostPerUser", 12000.00, 571429.0, "Total Cloud COGS / MAU"),
    ("CostPerTransaction", 3500.00, 2500000.0, "Transaction-related cost / transactions"),
    ("CostPer1KInputTokens", 7200.00, 120000.0, "LLM input cost / (input tokens / 1000)"),
    ("CostPer1KOutputTokens", 10800.00, 30000.0, "LLM output cost / (output tokens / 1000)"),
    ("CostPerRequest", 50000.00, 2000000.0, "AI total cost / total requests"),
]

# L5 预算预测样例：12 个月数据（Month 使用 1-12 数值，便于公式计算）
L5_HEADERS = ["Month", "ActualCost", "Budget", "MoM Growth%", "Forecast (Linear)", "Forecast (MA)"]
L5_MONTHS: List[int] = list(range(1, 13))
L5_ACTUAL_COSTS: List[float] = [
    12500.00, 13200.00, 14100.00, 13800.00, 15200.00, 15800.00,
    16400.00, 17200.00, 16900.00, 18100.00, 18800.00, 19500.00,
]
L5_BUDGETS: List[float] = [
    13000.00, 13500.00, 14000.00, 14500.00, 15000.00, 15500.00,
    16000.00, 16500.00, 17000.00, 17500.00, 18000.00, 18500.00,
]

# L6 承诺折扣优化样例
L6_HEADERS = [
    "Workload",
    "Interruptible",
    "OnDemandAnnual",
    "RI Annual",
    "SP Annual",
    "Spot Annual",
    "Recommended",
    "Savings",
]
L6_SAMPLE_DATA: List[Tuple[str, str, float, float, float, float]] = [
    ("WebFrontend", "No", 120000.00, 84000.00, 90000.00, 48000.00),
    ("BatchETL", "Yes", 80000.00, 56000.00, 60000.00, 32000.00),
    ("AIModelInference", "No", 240000.00, 168000.00, 180000.00, 120000.00),
    ("DataPipeline", "Yes", 60000.00, 42000.00, 45000.00, 24000.00),
    ("CacheCluster", "No", 45000.00, 31500.00, 33750.00, 18000.00),
    ("SearchIndexer", "Yes", 90000.00, 63000.00, 67500.00, 36000.00),
]

# L7 异常检测样例：30 条记录，其中 3 条为异常值
L7_HEADERS = ["Date", "Resource", "Service", "Cost", "Mean", "StdDev", "Z-Score", "IsAnomaly"]
L7_SAMPLE_DATA: List[Tuple[str, str, str, float]] = [
    ("2026-01-01", "resource-01", "EC2", 98.50),
    ("2026-01-02", "resource-02", "RDS", 102.30),
    ("2026-01-03", "resource-03", "S3", 95.00),
    ("2026-01-04", "resource-04", "Lambda", 105.20),
    ("2026-01-05", "resource-05", "EC2", 99.80),
    ("2026-01-06", "resource-06", "AI", 101.50),
    ("2026-01-07", "resource-07", "RDS", 97.20),
    ("2026-01-08", "resource-08", "EC2", 500.00),   # anomaly
    ("2026-01-09", "resource-09", "S3", 100.10),
    ("2026-01-10", "resource-10", "Lambda", 96.70),
    ("2026-01-11", "resource-11", "EC2", 103.40),
    ("2026-01-12", "resource-12", "RDS", 98.90),
    ("2026-01-13", "resource-13", "AI", 104.60),
    ("2026-01-14", "resource-14", "S3", 97.50),
    ("2026-01-15", "resource-15", "Lambda", 101.20),
    ("2026-01-16", "resource-16", "EC2", 99.30),
    ("2026-01-17", "resource-17", "RDS", 100.80),
    ("2026-01-18", "resource-18", "AI", 650.00),   # anomaly
    ("2026-01-19", "resource-19", "S3", 96.40),
    ("2026-01-20", "resource-20", "Lambda", 102.70),
    ("2026-01-21", "resource-21", "EC2", 98.10),
    ("2026-01-22", "resource-22", "RDS", 101.90),
    ("2026-01-23", "resource-23", "AI", 97.80),
    ("2026-01-24", "resource-24", "S3", 100.50),
    ("2026-01-25", "resource-25", "Lambda", 99.60),
    ("2026-01-26", "resource-26", "EC2", 103.10),
    ("2026-01-27", "resource-27", "RDS", 98.40),
    ("2026-01-28", "resource-28", "AI", 720.00),   # anomaly
    ("2026-01-29", "resource-29", "S3", 101.30),
    ("2026-01-30", "resource-30", "Lambda", 97.70),
]

# L8 AI GPU 分摊样例
L8_HEADERS = [
    "Team",
    "Project",
    "Model",
    "GPU Hours",
    "Token Input",
    "Token Output",
    "GPU Cost",
    "Token Cost",
    "Allocated Cost",
    "%Share",
]
L8_SAMPLE_DATA: List[Tuple[str, str, str, float, float, float, float, float]] = [
    ("AI-Platform", "Chatbot", "GPT-4", 500.0, 2_000_000.0, 500_000.0, 25000.00, 8000.00),
    ("Search-Team", "SemanticSearch", "Embedding-v3", 300.0, 800_000.0, 200_000.0, 15000.00, 3200.00),
    ("Data-Science", "ForecastModel", "LLaMA-3", 200.0, 500_000.0, 150_000.0, 10000.00, 1800.00),
]


# ---------------------------------------------------------------------------
# 样式辅助
# ---------------------------------------------------------------------------
def _make_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return header_font, header_fill, header_align, thin_border


def _style_header(ws, row: int, cols: int):
    header_font, header_fill, header_align, thin_border = _make_styles()
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_length:
                    max_length = val_len
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width


# ---------------------------------------------------------------------------
# 工作表构建
# ---------------------------------------------------------------------------
def _build_l1_raw_cost(ws):
    """L1-原始成本：云服务原始账单。"""
    ws.title = "L1-原始成本"
    ws.append(L1_HEADERS)
    _style_header(ws, 1, len(L1_HEADERS))

    for row in L1_SAMPLE_DATA:
        ws.append(row)

    # 设置数字格式
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = '#,##0.00'

    _auto_width(ws)


def _build_l2_rules(ws):
    """L2-分摊规则：分摊键定义。"""
    ws.title = "L2-分摊规则"
    ws.append(L2_HEADERS)
    _style_header(ws, 1, len(L2_HEADERS))

    for team, key, desc in L2_SAMPLE_RULES:
        ws.append([team, key, desc])

    # 校验合计为 100%
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
    ws.cell(row=total_row, column=2).number_format = '0.00%'
    for col_idx in range(1, len(L2_HEADERS) + 1):
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)

    # 百分比格式
    for row_idx in range(2, total_row):
        ws.cell(row=row_idx, column=2).number_format = '0.00%'

    _auto_width(ws)


def _build_l3_allocation(ws):
    """L3-分摊结果：使用 SUMIFS + XLOOKUP/VLOOKUP 分摊成本。"""
    ws.title = "L3-分摊结果"
    headers = ["Team", "DirectCost", "SharedCost", "AllocatedCost", "AllocationKey%", "RuleDescription"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    l1_sheet_name = "'L1-原始成本'"
    l2_sheet_name = "'L2-分摊规则'"

    for idx, (team, key, desc) in enumerate(L2_SAMPLE_RULES, start=2):
        # DirectCost: SUMIFS L1 Cost where Tag:Team == team
        direct_formula = (
            f"=SUMIFS({l1_sheet_name}!$C$2:$C$11,"
            f"{l1_sheet_name}!$D$2:$D$11,B{idx})"
        )

        # SharedCost: 将 Tag:Team="Platform" 的共享平台成本池按 L2 比例拆分
        shared_pool_formula = (
            f"SUMIFS({l1_sheet_name}!$C$2:$C$11,"
            f"{l1_sheet_name}!$D$2:$D$11,\"Platform\")"
        )
        shared_cost_formula = f"={shared_pool_formula}*E{idx}"

        # AllocatedCost = DirectCost + SharedCost
        allocated_formula = f"=B{idx}+C{idx}"

        # AllocationKey%: VLOOKUP/XLOOKUP from L2
        if _xlookup_available():
            key_formula = f"=XLOOKUP(B{idx},{l2_sheet_name}!$A$2:$A$5,{l2_sheet_name}!$B$2:$B$5,0)"
            desc_formula = f"=XLOOKUP(B{idx},{l2_sheet_name}!$A$2:$A$5,{l2_sheet_name}!$C$2:$C$5,\"\")"
        else:
            key_formula = f"=IFERROR(VLOOKUP(B{idx},{l2_sheet_name}!$A$2:$C$5,2,FALSE),0)"
            desc_formula = f"=IFERROR(VLOOKUP(B{idx},{l2_sheet_name}!$A$2:$C$5,3,FALSE),\"\")"

        ws.append([team, direct_formula, shared_cost_formula, allocated_formula, key_formula, desc_formula])

    # 合计行
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="合计")
    for col_idx in range(2, 5):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
        ws.cell(row=total_row, column=col_idx).number_format = '#,##0.00'
    ws.cell(row=total_row, column=5, value="=SUM(E2:E5)")
    ws.cell(row=total_row, column=5).number_format = '0.00%'
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)

    # 数字/百分比格式
    for row_idx in range(2, total_row):
        ws.cell(row=row_idx, column=2).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=3).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=5).number_format = '0.00%'

    _auto_width(ws)


def _xlookup_available() -> bool:
    """XLOOKUP 仅在较新 Excel 中可用；这里默认使用 VLOOKUP 保证兼容性。"""
    return False


def _build_l4_unit_economics(ws):
    """L4-单位经济学：每用户/每交易/每 Token 成本计算。"""
    ws.title = "L4-单位经济学"
    ws.append(L4_HEADERS)
    _style_header(ws, 1, len(L4_HEADERS))

    for idx, (metric, total_cost, unit_volume, notes) in enumerate(L4_SAMPLE_DATA, start=2):
        cost_per_unit_formula = f"=B{idx}/C{idx}"
        ws.append([metric, total_cost, unit_volume, cost_per_unit_formula, notes])

    # 合计/校验行
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TotalCloudCOGS")
    ws.cell(row=total_row, column=2, value="='L3-分摊结果'!D6")
    ws.cell(row=total_row, column=2).number_format = '#,##0.00'
    ws.cell(row=total_row, column=5, value="从 L3 分摊结果合计引用")
    for col_idx in range(1, len(L4_HEADERS) + 1):
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)

    # 数字格式
    for row_idx in range(2, total_row):
        ws.cell(row=row_idx, column=2).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=3).number_format = '#,##0'
        ws.cell(row=row_idx, column=4).number_format = '#,##0.000000'

    _auto_width(ws)


def _build_l5_budget_forecast(ws):
    """L5-预算预测：12 个月实际/预算、MoM 增长率、线性预测与移动平均。"""
    ws.title = "L5-预算预测"
    ws.append(L5_HEADERS)
    _style_header(ws, 1, len(L5_HEADERS))

    data_start = 2
    data_end = data_start + len(L5_MONTHS) - 1

    for idx, month in enumerate(L5_MONTHS, start=data_start):
        ws.append([month, L5_ACTUAL_COSTS[idx - data_start], L5_BUDGETS[idx - data_start], None, None, None])

    # 动态写入公式
    for row_idx in range(data_start, data_end + 1):
        # MoM Growth%（首月无上月，留空）
        if row_idx > data_start:
            ws.cell(row=row_idx, column=4, value=f"=(B{row_idx}-B{row_idx - 1})/B{row_idx - 1}")
        # Forecast (Linear)：基于 12 个月实际成本做线性回归预测
        ws.cell(
            row=row_idx,
            column=5,
            value=f"=FORECAST.LINEAR(A{row_idx},$B${data_start}:$B${data_end},$A${data_start}:$A${data_end})",
        )
        # Forecast (MA)：近 3 个月移动平均（前两个月数据不足，留空）
        if row_idx >= data_start + 2:
            ws.cell(row=row_idx, column=6, value=f"=AVERAGE(B{row_idx - 2}:B{row_idx})")

    # 汇总行
    summary_row = data_end + 2
    ws.cell(row=summary_row, column=1, value="年度总成本")
    ws.cell(row=summary_row, column=2, value=f"=SUM(B{data_start}:B{data_end})")

    ws.cell(row=summary_row + 1, column=1, value="年度预算")
    ws.cell(row=summary_row + 1, column=2, value=f"=SUM(C{data_start}:C{data_end})")

    ws.cell(row=summary_row + 2, column=1, value="偏差")
    ws.cell(row=summary_row + 2, column=2, value=f"=B{summary_row}-B{summary_row + 1}")

    ws.cell(row=summary_row + 3, column=1, value="执行率")
    ws.cell(row=summary_row + 3, column=2, value=f"=B{summary_row}/B{summary_row + 1}")

    for row_idx in range(summary_row, summary_row + 4):
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

    # 数字格式
    for row_idx in range(data_start, data_end + 1):
        ws.cell(row=row_idx, column=2).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=3).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=4).number_format = '0.00%'
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00'

    for row_idx in range(summary_row, summary_row + 4):
        ws.cell(row=row_idx, column=2).number_format = '#,##0.00'
    ws.cell(row=summary_row + 3, column=2).number_format = '0.00%'

    _auto_width(ws)


def _build_l6_commitment_discount(ws):
    """L6-承诺折扣优化：对比按需/RI/SP/Spot，推荐最低年度成本。"""
    ws.title = "L6-承诺折扣优化"
    ws.append(L6_HEADERS)
    _style_header(ws, 1, len(L6_HEADERS))

    data_start = 2
    for idx, (workload, interruptible, on_demand, ri, sp, spot) in enumerate(L6_SAMPLE_DATA, start=data_start):
        ws.append([workload, interruptible, on_demand, ri, sp, spot, None, None])
        # Recommended：可中断时考虑 Spot，否则仅比较按需/RI/SP
        ws.cell(
            row=idx,
            column=7,
            value=f'=IF(B{idx}="Yes",MIN(C{idx}:F{idx}),MIN(C{idx}:E{idx}))',
        )
        # Savings = OnDemand - Recommended
        ws.cell(row=idx, column=8, value=f"=C{idx}-G{idx}")

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=3, value=f"=SUM(C{data_start}:C{total_row - 1})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G{data_start}:G{total_row - 1})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H{data_start}:H{total_row - 1})")

    savings_rate_row = total_row + 1
    ws.cell(row=savings_rate_row, column=1, value="节省率")
    ws.cell(row=savings_rate_row, column=8, value=f"=H{total_row}/C{total_row}")

    for row_idx in (total_row, savings_rate_row):
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

    # 数字格式
    for row_idx in range(data_start, total_row):
        for col_idx in (3, 4, 5, 6, 7, 8):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'
    for col_idx in (3, 7, 8):
        ws.cell(row=total_row, column=col_idx).number_format = '#,##0.00'
    ws.cell(row=savings_rate_row, column=8).number_format = '0.00%'

    _auto_width(ws)


def _build_l7_anomaly_detection(ws):
    """L7-异常检测：使用均值、标准差与 Z-Score 标记异常成本。"""
    ws.title = "L7-异常检测"
    ws.append(L7_HEADERS)
    _style_header(ws, 1, len(L7_HEADERS))

    data_start = 2
    for idx, (date, resource, service, cost) in enumerate(L7_SAMPLE_DATA, start=data_start):
        ws.append([date, resource, service, cost, None, None, None, None])

    data_end = ws.max_row

    for row_idx in range(data_start, data_end + 1):
        # Mean：整列平均
        ws.cell(row=row_idx, column=5, value=f"=AVERAGE($D${data_start}:$D${data_end})")
        # StdDev：样本标准差
        ws.cell(row=row_idx, column=6, value=f"=STDEV.S($D${data_start}:$D${data_end})")
        # Z-Score
        ws.cell(row=row_idx, column=7, value=f"=(D{row_idx}-E{row_idx})/F{row_idx}")
        # IsAnomaly
        ws.cell(row=row_idx, column=8, value=f'=IF(ABS(G{row_idx})>2,"Yes","No")')

    # 汇总
    summary_row = data_end + 2
    ws.cell(row=summary_row, column=1, value="异常数")
    ws.cell(row=summary_row, column=8, value=f'=COUNTIF($H${data_start}:$H${data_end},"Yes")')
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    # 数字格式
    for row_idx in range(data_start, data_end + 1):
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=7).number_format = '0.00'

    _auto_width(ws)


def _build_l8_ai_gpu_allocation(ws):
    """L8-AI GPU 分摊：按团队/项目/模型分摊 GPU 与 Token 成本。"""
    ws.title = "L8-AI GPU 分摊"
    ws.append(L8_HEADERS)
    _style_header(ws, 1, len(L8_HEADERS))

    data_start = 2
    for idx, (team, project, model, gpu_hours, token_in, token_out, gpu_cost, token_cost) in enumerate(
        L8_SAMPLE_DATA, start=data_start
    ):
        ws.append([team, project, model, gpu_hours, token_in, token_out, gpu_cost, token_cost, None, None])
        # Allocated Cost = GPU Cost + Token Cost
        ws.cell(row=idx, column=9, value=f"=G{idx}+H{idx}")
        # %Share
        ws.cell(row=idx, column=10, value=f"=I{idx}/SUM($I${data_start}:$I${data_start + len(L8_SAMPLE_DATA) - 1})")

    data_end = ws.max_row

    # 汇总
    total_row = data_end + 2
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=9, value=f"=SUM(I{data_start}:I{data_end})")
    ws.cell(row=total_row, column=10, value=f"=SUM(J{data_start}:J{data_end})")
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    # 数字格式
    for row_idx in range(data_start, data_end + 1):
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=5).number_format = '#,##0'
        ws.cell(row=row_idx, column=6).number_format = '#,##0'
        ws.cell(row=row_idx, column=7).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=8).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=9).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=10).number_format = '0.00%'

    ws.cell(row=total_row, column=9).number_format = '#,##0.00'
    ws.cell(row=total_row, column=10).number_format = '0.00%'

    _auto_width(ws)


def _build_readme_sheet(ws):
    """额外增加一个使用说明工作表。"""
    ws.title = "使用说明"
    instructions = [
        ["FinOps L1–L8 成本分摊 Excel 公式模板", ""],
        ["", ""],
        ["工作表", "说明"],
        ["L1-原始成本", "填写或粘贴云服务原始账单；Tag:Team 用于直接归属。"],
        ["L2-分摊规则", "维护 Team 的分摊键，合计应等于 100%。"],
        ["L3-分摊结果", "自动按 SUMIFS 汇总直接成本，并按 L2 比例拆分共享成本。"],
        ["L4-单位经济学", "基于 L3 的 TotalCloudCOGS 或独立成本池计算单位成本。"],
        ["L5-预算预测", "12 个月 ActualCost 与 Budget，含 MoM Growth%、FORECAST.LINEAR、3 个月移动平均及年度执行率。"],
        ["L6-承诺折扣优化", "对比 OnDemand / RI / SP / Spot 年度成本，自动推荐最低方案并计算 Savings。"],
        ["L7-异常检测", "基于 AVERAGE、STDEV.S 与 Z-Score（阈值 2）自动标记异常成本记录。"],
        ["L8-AI GPU 分摊", "按 Team/Project/Model 汇总 GPU Hours、Token 与成本，并计算 Allocated Cost 占比。"],
        ["", ""],
        ["使用提示", ""],
        ["1", "将 L1 替换为贵组织的真实云账单（保持列结构一致）。"],
        ["2", "在 L2 中调整 AllocationKey%，确保团队比例符合 FinOps 分摊策略。"],
        ["3", "L3 中的公式会自动重新计算；如需更复杂的共享池逻辑，可拆分 Platform/Data 等共享项。"],
        ["4", "L4 中的单位经济学公式可按业务指标（MAU、交易数、Token 数）自行扩展。"],
        ["5", "L5–L8 中的公式示例可直接替换为真实数据，预测与异常阈值可按业务需求调整。"],
        ["", ""],
        ["对齐来源", ""],
        ["FinOps Foundation Framework 2026", ""],
        ["FOCUS (FinOps Open Cost and Usage Specification)", ""],
    ]
    for row in instructions:
        ws.append(row)

    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=3, column=2).font = Font(bold=True)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 90


# ---------------------------------------------------------------------------
# 主生成逻辑
# ---------------------------------------------------------------------------
def generate_template(output_path: Path) -> None:
    if not HAS_OPENPYXL:
        print(
            "错误：未检测到 openpyxl。请运行以下命令安装：\n"
            "  pip install openpyxl\n"
            "或使用虚拟环境：\n"
            "  python -m venv .venv\n"
            "  source .venv/bin/activate  # Windows: .venv\\Scripts\activate\n"
            "  pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # 删除默认 sheet，按顺序创建
    wb.remove(wb.active)

    ws_readme = wb.create_sheet()
    _build_readme_sheet(ws_readme)

    ws_l1 = wb.create_sheet()
    _build_l1_raw_cost(ws_l1)

    ws_l2 = wb.create_sheet()
    _build_l2_rules(ws_l2)

    ws_l3 = wb.create_sheet()
    _build_l3_allocation(ws_l3)

    ws_l4 = wb.create_sheet()
    _build_l4_unit_economics(ws_l4)

    ws_l5 = wb.create_sheet()
    _build_l5_budget_forecast(ws_l5)

    ws_l6 = wb.create_sheet()
    _build_l6_commitment_discount(ws_l6)

    ws_l7 = wb.create_sheet()
    _build_l7_anomaly_detection(ws_l7)

    ws_l8 = wb.create_sheet()
    _build_l8_ai_gpu_allocation(ws_l8)

    wb.save(str(output_path))
    print(f"FinOps Excel 公式模板已生成: {output_path.resolve()}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="FinOps Excel 公式模板生成器 (L1–L8 Cost Allocation Template Generator)"
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"输出文件路径（默认: {DEFAULT_OUTPUT}）",
    )
    args = parser.parse_args()

    try:
        generate_template(args.output)
    except Exception as exc:
        print(f"错误: 生成模板失败: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
