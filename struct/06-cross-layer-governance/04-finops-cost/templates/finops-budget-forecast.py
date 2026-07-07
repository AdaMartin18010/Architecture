#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FinOps 预算与预测分析工具 (FinOps Budget & Forecast Analyzer)

功能：
    1. 读取 YAML / JSON / CSV 格式的历史成本数据（至少支持 12 个月）。
    2. 计算以下核心指标：
       - 月度总成本 (Monthly Total Cost)
       - 月度增长率 MoM (Month-over-Month Growth Rate)
       - 年度总成本 (Annual Total Cost)
       - 下季度 / 下半年预测（简单线性回归 + 3 个月移动平均）
       - 预算偏差 (Budget Variance: Actual - Budget)
       - 预算执行率 / 运行率 Run Rate (Annualized Actual / Annual Budget)
    3. 输出三种格式：
       - 控制台表格
       - CSV 报告（默认 reports/finops-budget-forecast.csv）
       - Excel 报告（openpyxl 可用时），含历史数据、预测、预算偏差三个工作表

用法：
    python finops-budget-forecast.py --input example-budget.yaml --format console
    python finops-budget-forecast.py --input costs.csv --budget 500000 --format csv
    python finops-budget-forecast.py --input costs.json --budget 600000 --output forecast.xlsx --format xlsx

依赖：
    - 标准库：argparse, csv, json, pathlib, statistics, sys, typing
    - 可选：PyYAML（解析 YAML，缺失时仅支持 JSON/CSV）
    - 可选：openpyxl（生成 .xlsx，缺失时降级为 CSV）

对齐来源：
    - FinOps Foundation Forecasting Capability
      https://www.finops.org/framework/capabilities/forecasting/
    - FOCUS 1.0 (FinOps Open Cost and Usage Specification)
      https://focus.finops.org/

本文件为 Python 工具脚本，不参与 Markdown 质量门控。
"""

from __future__ import annotations

import argparse
import csv
import json
import statistics
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

# ---------------------------------------------------------------------------
# 可选依赖：优雅降级
# ---------------------------------------------------------------------------
try:
    import yaml  # type: ignore

    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
DEFAULT_OUTPUT = Path("reports/finops-budget-forecast.csv")
MIN_HISTORY_MONTHS = 12


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------
def _load_yaml(path: Path) -> Dict[str, Any]:
    if not HAS_YAML:
        raise RuntimeError(
            "PyYAML is required to parse YAML input. "
            "Install it with: pip install pyyaml"
        )
    with path.open("r", encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def _load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _load_csv(path: Path) -> Dict[str, Any]:
    """CSV 支持两种格式：
    1) period,cost,budget（单个月份一行）
    2) period,cost（预算通过 --budget 传入）
    """
    history: List[Dict[str, Any]] = []
    currency = "USD"
    budget_annual: Optional[float] = None

    with path.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            period = (row.get("period") or row.get("month") or "").strip()
            cost_raw = row.get("cost") or row.get("amount") or "0"
            try:
                cost = float(cost_raw)
            except ValueError:
                continue
            entry: Dict[str, Any] = {"period": period, "cost": cost}
            if "budget" in row and row["budget"]:
                entry["budget"] = float(row["budget"])
            history.append(entry)

    # 尝试从 CSV 推断年度预算：取最近 12 个月 budget 列之和
    budget_values = [h.get("budget") for h in history if isinstance(h.get("budget"), (int, float))]
    if len(budget_values) >= MIN_HISTORY_MONTHS:
        budget_annual = sum(budget_values[-MIN_HISTORY_MONTHS:])

    return {
        "metadata": {"currency": currency, "budget_annual": budget_annual},
        "history": history,
    }


def load_input(path: Path) -> Dict[str, Any]:
    """根据扩展名自动选择加载器。"""
    suffix = path.suffix.lower()
    if suffix in (".yaml", ".yml"):
        return _load_yaml(path)
    if suffix == ".json":
        return _load_json(path)
    if suffix == ".csv":
        return _load_csv(path)
    # 尝试自动探测
    with path.open("r", encoding="utf-8") as fh:
        first = fh.read(1)
    if first in ("{", "["):
        return _load_json(path)
    if HAS_YAML:
        return _load_yaml(path)
    return _load_csv(path)


# ---------------------------------------------------------------------------
# 计算逻辑
# ---------------------------------------------------------------------------
def _linear_regression(xs: List[float], ys: List[float]) -> Tuple[float, float]:
    """最小二乘法线性回归：返回 (slope, intercept)。"""
    n = len(xs)
    if n < 2:
        return 0.0, statistics.mean(ys) if ys else 0.0
    mean_x = statistics.mean(xs)
    mean_y = statistics.mean(ys)
    numerator = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    denominator = sum((x - mean_x) ** 2 for x in xs)
    slope = numerator / denominator if denominator != 0 else 0.0
    intercept = mean_y - slope * mean_x
    return slope, intercept


def _forecast_linear(history: List[Dict[str, Any]], periods: int) -> List[Dict[str, Any]]:
    """基于简单线性回归预测未来 periods 个月。"""
    xs = list(range(len(history)))
    ys = [h["cost"] for h in history]
    slope, intercept = _linear_regression(xs, ys)
    last_period = history[-1]["period"]
    base_date = _parse_period(last_period)
    result: List[Dict[str, Any]] = []
    for i in range(1, periods + 1):
        idx = len(history) - 1 + i
        predicted = slope * idx + intercept
        predicted = max(predicted, 0.0)  # 成本不为负
        period = _next_period(base_date, i)
        result.append({
            "period": period,
            "cost": round(predicted, 2),
            "type": "forecast-linear",
        })
    return result


def _forecast_moving_average(history: List[Dict[str, Any]], window: int, periods: int) -> List[Dict[str, Any]]:
    """基于最近 window 个月移动平均预测未来 periods 个月。"""
    if len(history) < window:
        window = len(history)
    recent = [h["cost"] for h in history[-window:]]
    avg = statistics.mean(recent)
    last_period = history[-1]["period"]
    base_date = _parse_period(last_period)
    result: List[Dict[str, Any]] = []
    for i in range(1, periods + 1):
        period = _next_period(base_date, i)
        result.append({
            "period": period,
            "cost": round(avg, 2),
            "type": "forecast-ma",
        })
    return result


def _parse_period(period: str) -> Tuple[int, int]:
    """解析 YYYY-MM 格式。"""
    parts = period.replace("-", "/").split("/")
    year = int(parts[0])
    month = int(parts[1])
    return year, month


def _next_period(base: Tuple[int, int], offset: int) -> str:
    """计算 base 月份后 offset 个月的 YYYY-MM。"""
    year, month = base
    total_months = year * 12 + (month - 1) + offset
    new_year = total_months // 12
    new_month = total_months % 12 + 1
    return f"{new_year:04d}-{new_month:02d}"


def compute_metrics(history: List[Dict[str, Any]], annual_budget: float) -> Dict[str, Any]:
    """计算全部指标。"""
    if len(history) < MIN_HISTORY_MONTHS:
        raise ValueError(
            f"历史数据至少需要 {MIN_HISTORY_MONTHS} 个月，当前仅 {len(history)} 个月。"
        )

    # 按时间排序并保证 cost 为 float
    history = sorted(history, key=lambda h: h.get("period", ""))
    for h in history:
        h["cost"] = float(h.get("cost", 0.0))
        h["type"] = "actual"

    # 月度 MoM 增长率
    for i, h in enumerate(history):
        if i == 0:
            h["mom_growth"] = None
        else:
            prev = history[i - 1]["cost"]
            if prev == 0:
                h["mom_growth"] = None
            else:
                h["mom_growth"] = round((h["cost"] - prev) / prev * 100, 2)

    # 年度总成本（最近 12 个月）
    annual_total = sum(h["cost"] for h in history[-MIN_HISTORY_MONTHS:])

    # 预测：下季度（3 个月）和下半年（6 个月）
    forecast_quarter = _forecast_linear(history, 3)
    forecast_half = _forecast_moving_average(history, 3, 6)

    # 年度化运行率 (Annualized Run Rate)
    run_rate = annual_total  # 最近 12 个月即年化

    # 预算偏差与执行率
    budget_variance = annual_total - annual_budget
    budget_execution_rate = (run_rate / annual_budget * 100) if annual_budget else 0.0

    return {
        "history": history,
        "forecast_quarter": forecast_quarter,
        "forecast_half": forecast_half,
        "annual_total": round(annual_total, 2),
        "annual_budget": round(annual_budget, 2),
        "budget_variance": round(budget_variance, 2),
        "budget_execution_rate": round(budget_execution_rate, 2),
        "run_rate": round(run_rate, 2),
    }


# ---------------------------------------------------------------------------
# 输出
# ---------------------------------------------------------------------------
def _fmt_money(value: Optional[float], currency: str = "USD") -> str:
    if value is None:
        return "N/A"
    return f"{currency} {value:,.2f}"


def _fmt_pct(value: Optional[float]) -> str:
    if value is None:
        return "N/A"
    return f"{value:.2f}%"


def output_console(metrics: Dict[str, Any], currency: str) -> None:
    """在控制台打印历史数据、预测与预算偏差表格。"""
    print("=" * 80)
    print("FinOps 预算与预测报告")
    print("=" * 80)
    print(f"\n币种: {currency}")
    print(f"最近 12 个月总成本: {_fmt_money(metrics['annual_total'], currency)}")
    print(f"年度预算:           {_fmt_money(metrics['annual_budget'], currency)}")
    print(f"预算偏差:           {_fmt_money(metrics['budget_variance'], currency)}")
    print(f"预算执行率 (Run Rate): {_fmt_pct(metrics['budget_execution_rate'])}")
    print()

    # 历史数据表
    print("-" * 80)
    print("历史月度成本")
    print("-" * 80)
    print(f"{'Period':<12}{'Cost':>16}{'MoM Growth':>16}{'Type':>14}")
    print("-" * 80)
    for h in metrics["history"]:
        print(
            f"{h['period']:<12}"
            f"{_fmt_money(h['cost'], currency):>16}"
            f"{_fmt_pct(h.get('mom_growth')):>16}"
            f"{h.get('type', 'actual'):>14}"
        )

    # 预测表
    print("\n" + "-" * 80)
    print("下季度预测（线性回归）")
    print("-" * 80)
    print(f"{'Period':<12}{'Forecast':>16}{'Method':>20}")
    print("-" * 80)
    for f in metrics["forecast_quarter"]:
        print(f"{f['period']:<12}{_fmt_money(f['cost'], currency):>16}{f['type']:>20}")

    print("\n" + "-" * 80)
    print("下半年预测（3 个月移动平均）")
    print("-" * 80)
    print(f"{'Period':<12}{'Forecast':>16}{'Method':>20}")
    print("-" * 80)
    for f in metrics["forecast_half"]:
        print(f"{f['period']:<12}{_fmt_money(f['cost'], currency):>16}{f['type']:>20}")

    print("\n" + "=" * 80)


def output_csv(metrics: Dict[str, Any], output_path: Path, currency: str) -> None:
    """导出 CSV 报告。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Currency", currency])
        writer.writerow(["AnnualTotal", metrics["annual_total"]])
        writer.writerow(["AnnualBudget", metrics["annual_budget"]])
        writer.writerow(["BudgetVariance", metrics["budget_variance"]])
        writer.writerow(["BudgetExecutionRate", f"{metrics['budget_execution_rate']}%"])
        writer.writerow(["RunRate", metrics["run_rate"]])
        writer.writerow([])

        writer.writerow(["Period", "Cost", "MoMGrowth", "Type"])
        for h in metrics["history"]:
            writer.writerow([
                h["period"],
                h["cost"],
                f"{h['mom_growth']}%" if h.get("mom_growth") is not None else "",
                h.get("type", "actual"),
            ])
        writer.writerow([])

        writer.writerow(["ForecastPeriod", "ForecastCost", "Method"])
        for f in metrics["forecast_quarter"]:
            writer.writerow([f["period"], f["cost"], f"linear-regression"])
        for f in metrics["forecast_half"]:
            writer.writerow([f["period"], f["cost"], f"moving-average-3m"])

    print(f"CSV 报告已保存: {output_path}")


def _style_header_row(ws, row: int, cols: int) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _write_sheet(ws, headers: List[str], rows: Iterable[List[Any]]) -> None:
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for row in rows:
        ws.append(row)
    # 自动调整列宽
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)


def output_excel(metrics: Dict[str, Any], output_path: Path, currency: str) -> None:
    """导出 Excel 报告，包含三个工作表。"""
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is required for Excel output. "
            "Install it with: pip install openpyxl"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # 工作表 1：历史数据
    ws_history = wb.active
    ws_history.title = "历史数据"
    _write_sheet(
        ws_history,
        ["Period", "Cost", "MoM Growth (%)", "Type"],
        [
            [h["period"], h["cost"], h.get("mom_growth"), h.get("type", "actual")]
            for h in metrics["history"]
        ],
    )

    # 工作表 2：预测
    ws_forecast = wb.create_sheet("预测")
    forecast_rows: List[List[Any]] = []
    for f in metrics["forecast_quarter"]:
        forecast_rows.append([f["period"], f["cost"], "linear-regression"])
    for f in metrics["forecast_half"]:
        forecast_rows.append([f["period"], f["cost"], "moving-average-3m"])
    _write_sheet(ws_forecast, ["Period", "Forecast Cost", "Method"], forecast_rows)

    # 工作表 3：预算偏差
    ws_budget = wb.create_sheet("预算偏差")
    _write_sheet(
        ws_budget,
        ["Metric", "Value"],
        [
            ["Currency", currency],
            ["Annual Total", metrics["annual_total"]],
            ["Annual Budget", metrics["annual_budget"]],
            ["Budget Variance", metrics["budget_variance"]],
            ["Budget Execution Rate (%)", metrics["budget_execution_rate"]],
            ["Run Rate", metrics["run_rate"]],
        ],
    )

    wb.save(output_path)
    print(f"Excel 报告已保存: {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="FinOps 预算与预测分析工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="输入文件路径（YAML/JSON/CSV）",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"输出文件路径（默认 {DEFAULT_OUTPUT}）",
    )
    parser.add_argument(
        "--budget",
        type=float,
        default=None,
        help="年度预算（覆盖输入文件中的 budget_annual）",
    )
    parser.add_argument(
        "--format",
        choices=["console", "csv", "xlsx"],
        default="console",
        help="输出格式：console / csv / xlsx（默认 console）",
    )
    return parser


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if not args.input.exists():
        print(f"错误：输入文件不存在 {args.input}", file=sys.stderr)
        return 1

    raw = load_input(args.input)
    metadata = raw.get("metadata") or {}
    history = raw.get("history") or []

    if not history:
        print("错误：输入文件未包含历史成本数据（history）。", file=sys.stderr)
        return 1

    # 预算优先级：CLI > 文件 budget_annual > 文件 monthly_budget * 12 > CSV 推断
    annual_budget: Optional[float] = args.budget
    if annual_budget is None:
        annual_budget = metadata.get("budget_annual")
    if annual_budget is None and metadata.get("monthly_budget"):
        annual_budget = float(metadata["monthly_budget"]) * 12

    if annual_budget is None:
        print(
            "错误：未提供年度预算。请使用 --budget 或在输入文件中指定 budget_annual/monthly_budget。",
            file=sys.stderr,
        )
        return 1

    currency = metadata.get("currency", "USD")

    try:
        metrics = compute_metrics(history, float(annual_budget))
    except ValueError as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 1

    if args.format == "console":
        output_console(metrics, currency)
    elif args.format == "csv":
        output_csv(metrics, args.output, currency)
    elif args.format == "xlsx":
        if not HAS_OPENPYXL:
            print(
                "警告：openpyxl 未安装，已降级为 CSV 输出。",
                file=sys.stderr,
            )
            output_csv(metrics, args.output.with_suffix(".csv"), currency)
        else:
            output_excel(metrics, args.output.with_suffix(".xlsx"), currency)
    else:
        parser.print_help()
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
