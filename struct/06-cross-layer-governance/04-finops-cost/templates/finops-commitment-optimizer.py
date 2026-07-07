#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FinOps 承诺折扣优化器 (Commitment Optimizer)

功能：
1. 读取 YAML/JSON 输入（按需成本、RI/SP/Spot 折扣率、工作负载可中断性）。
2. 计算三种场景的年度总成本：
   - 全按需 (On-Demand)
   - RI/SP + Spot 最优组合
   - 全 Spot（不可中断负载保持按需）
3. 输出推荐方案、预计年节省金额、节省百分比、风险等级。
4. 支持控制台、CSV、可选 Excel 三种输出。

对齐来源：
- FinOps Foundation Rate Optimization Capability
- AWS Reserved Instances / Savings Plans / Spot Instances 文档
- Azure Reserved VM Instances / Savings Plans / Spot VMs 文档
- Google Cloud Committed Use Discounts / Spot VMs 文档

约束：
- 标准库优先；openpyxl / PyYAML 缺失时优雅降级。
- 不硬编码任何云厂商价格，全部参数通过 YAML/JSON 输入注入。
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import asdict, dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import yaml  # type: ignore

    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font  # type: ignore

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def d(value: Any) -> Decimal:
    """将数值/字符串安全转换为 Decimal。"""
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value) if value is not None else "0")


def fmt_currency(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}"


def fmt_percent(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}%"


# ---------------------------------------------------------------------------
# 数据模型
# ---------------------------------------------------------------------------

@dataclass
class Discount:
    kind: str  # ri | savings_plans | spot
    discount_rate: Decimal
    commitment_years: int = 0
    upfront_ratio: Decimal = field(default_factory=lambda: Decimal("0"))
    interruption_risk_score: Optional[int] = None  # 仅 Spot，1-10


@dataclass
class Workload:
    workload_id: str
    name: str
    annual_on_demand_cost: Decimal
    can_interrupt: bool
    recommended_commitment_type: Optional[str] = None
    notes: str = ""


@dataclass
class ScenarioResult:
    name: str
    total_annual_cost: Decimal
    annual_savings: Decimal
    savings_percent: Decimal
    spot_share_percent: Decimal
    risk_class: str  # 低 | 中 | 高
    risk_label: str
    breakdown: List[Dict[str, Any]] = field(default_factory=list)


# ---------------------------------------------------------------------------
# 输入解析
# ---------------------------------------------------------------------------

def load_input(path: Path) -> Dict[str, Any]:
    """读取 YAML 或 JSON 输入；PyYAML 缺失时降级为 JSON。"""
    text = path.read_text(encoding="utf-8")
    suffix = path.suffix.lower()

    if suffix in (".yaml", ".yml"):
        if HAS_YAML:
            return yaml.safe_load(text)
        print("[警告] 未安装 PyYAML，无法解析 YAML，尝试按 JSON 解析...", file=sys.stderr)
        try:
            return json.loads(text)
        except json.JSONDecodeError as exc:
            raise RuntimeError(
                "当前环境缺少 PyYAML，且该 YAML 文件无法作为 JSON 解析。"
                "请安装 PyYAML：pip install pyyaml"
            ) from exc

    if suffix == ".json":
        return json.loads(text)

    # 未知后缀时优先尝试 YAML，再回退 JSON
    if HAS_YAML:
        try:
            return yaml.safe_load(text)
        except Exception:
            pass
    return json.loads(text)


def parse_discounts(raw: Dict[str, Any]) -> Dict[str, Discount]:
    """解析 RI / SP / Spot 折扣配置。"""
    discounts: Dict[str, Discount] = {}

    ri = raw.get("ri") or raw.get("reserved_instances")
    if ri:
        discounts["ri"] = Discount(
            kind="ri",
            discount_rate=d(ri.get("discount_rate", 0)),
            commitment_years=int(ri.get("commitment_years", 1)),
            upfront_ratio=d(ri.get("upfront_ratio", 0)),
        )

    sp = raw.get("savings_plans") or raw.get("sp")
    if sp:
        discounts["savings_plans"] = Discount(
            kind="savings_plans",
            discount_rate=d(sp.get("discount_rate", 0)),
            commitment_years=int(sp.get("commitment_years", 1)),
            upfront_ratio=d(sp.get("upfront_ratio", 0)),
        )

    spot = raw.get("spot")
    if spot:
        discounts["spot"] = Discount(
            kind="spot",
            discount_rate=d(spot.get("discount_rate", 0)),
            commitment_years=0,
            upfront_ratio=Decimal("0"),
            interruption_risk_score=int(spot.get("interruption_risk_score", 5)),
        )

    return discounts


def parse_workloads(raw_list: List[Dict[str, Any]]) -> List[Workload]:
    workloads: List[Workload] = []
    for idx, item in enumerate(raw_list):
        wid = item.get("id") or item.get("workload_id") or f"workload-{idx + 1}"
        name = item.get("name") or wid
        annual = item.get("annual_on_demand_cost") or item.get("annual_cost")
        if annual is None:
            monthly = item.get("monthly_on_demand_cost") or item.get("monthly_cost")
            annual = d(monthly) * 12 if monthly is not None else Decimal("0")
        workloads.append(
            Workload(
                workload_id=wid,
                name=name,
                annual_on_demand_cost=d(annual),
                can_interrupt=bool(item.get("can_interrupt", False)),
                recommended_commitment_type=item.get("recommended_commitment_type"),
                notes=item.get("notes", ""),
            )
        )
    return workloads


# ---------------------------------------------------------------------------
# 优化计算
# ---------------------------------------------------------------------------

def choose_commitment_discount(
    workload: Workload, discounts: Dict[str, Discount]
) -> Tuple[Optional[str], Decimal]:
    """
    为非中断型工作负载选择最优承诺折扣。
    优先级：工作负载显式推荐 > RI/SP 中折扣最高者。
    """
    ctype = (workload.recommended_commitment_type or "").lower()
    if ctype in ("ri", "reserved_instances") and "ri" in discounts:
        return "RI", discounts["ri"].discount_rate
    if ctype in ("sp", "savings_plans") and "savings_plans" in discounts:
        return "SP", discounts["savings_plans"].discount_rate

    best_rate = Decimal("0")
    best_label = "On-Demand"
    for key in ("ri", "savings_plans"):
        if key in discounts:
            if discounts[key].discount_rate > best_rate:
                best_rate = discounts[key].discount_rate
                best_label = "RI" if key == "ri" else "SP"
    return (best_label if best_rate > 0 else None), best_rate


def build_scenarios(
    workloads: List[Workload],
    discounts: Dict[str, Discount],
) -> List[ScenarioResult]:
    baseline = sum((w.annual_on_demand_cost for w in workloads), Decimal("0"))
    spot = discounts.get("spot")
    spot_rate = spot.discount_rate if spot else Decimal("0")
    spot_risk_score = spot.interruption_risk_score if spot else 5

    scenarios: List[ScenarioResult] = []

    # 1. 全按需
    on_demand_breakdown = []
    for w in workloads:
        on_demand_breakdown.append(
            {
                "workload_id": w.workload_id,
                "name": w.name,
                "can_interrupt": w.can_interrupt,
                "assigned_type": "On-Demand",
                "discount_rate": Decimal("0"),
                "annual_cost": w.annual_on_demand_cost,
            }
        )
    scenarios.append(
        _make_result(
            "全按需 (On-Demand)",
            baseline,
            baseline,
            on_demand_breakdown,
            spot_risk_score,
            force_risk_class="低",
        )
    )

    # 2. RI/SP + Spot 最优组合
    optimal_breakdown = []
    for w in workloads:
        if w.can_interrupt and spot:
            assigned = "Spot"
            rate = spot_rate
        else:
            assigned, rate = choose_commitment_discount(w, discounts)
            if assigned is None:
                assigned = "On-Demand"
                rate = Decimal("0")
        cost = w.annual_on_demand_cost * (Decimal("1") - rate)
        optimal_breakdown.append(
            {
                "workload_id": w.workload_id,
                "name": w.name,
                "can_interrupt": w.can_interrupt,
                "assigned_type": assigned,
                "discount_rate": rate,
                "annual_cost": cost,
            }
        )
    optimal_total = sum((row["annual_cost"] for row in optimal_breakdown), Decimal("0"))
    scenarios.append(
        _make_result(
            "RI/SP + Spot 最优组合",
            baseline,
            optimal_total,
            optimal_breakdown,
            spot_risk_score,
        )
    )

    # 3. 全 Spot（不可中断负载保持按需）
    all_spot_breakdown = []
    has_non_interruptible = False
    for w in workloads:
        if w.can_interrupt and spot:
            assigned = "Spot"
            rate = spot_rate
        else:
            has_non_interruptible = True
            assigned = "On-Demand（不可中断）"
            rate = Decimal("0")
        cost = w.annual_on_demand_cost * (Decimal("1") - rate)
        all_spot_breakdown.append(
            {
                "workload_id": w.workload_id,
                "name": w.name,
                "can_interrupt": w.can_interrupt,
                "assigned_type": assigned,
                "discount_rate": rate,
                "annual_cost": cost,
            }
        )
    all_spot_total = sum((row["annual_cost"] for row in all_spot_breakdown), Decimal("0"))
    risk_override = "高（含不可中断负载）" if has_non_interruptible else None
    scenarios.append(
        _make_result(
            "全 Spot（仅可中断负载适用）",
            baseline,
            all_spot_total,
            all_spot_breakdown,
            spot_risk_score,
            risk_label_override=risk_override,
        )
    )

    return scenarios


def _make_result(
    name: str,
    baseline: Decimal,
    total: Decimal,
    breakdown: List[Dict[str, Any]],
    spot_risk_score: int,
    force_risk_class: Optional[str] = None,
    risk_label_override: Optional[str] = None,
) -> ScenarioResult:
    annual_savings = baseline - total
    savings_percent = (
        (annual_savings / baseline * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if baseline > 0
        else Decimal("0")
    )

    spot_cost = sum(
        (row["annual_cost"] for row in breakdown if row["assigned_type"] == "Spot"),
        Decimal("0"),
    )
    spot_share_percent = (
        (spot_cost / total * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if total > 0
        else Decimal("0")
    )

    if force_risk_class:
        risk_class = force_risk_class
    else:
        # 风险得分：Spot 成本占比 与 Spot 中断风险评分取高
        score_from_share = spot_share_percent
        score_from_risk = Decimal(spot_risk_score) * Decimal("10")
        score = max(score_from_share, score_from_risk)
        if score <= Decimal("30"):
            risk_class = "低"
        elif score <= Decimal("60"):
            risk_class = "中"
        else:
            risk_class = "高"

    risk_label = risk_label_override or f"{risk_class}"

    return ScenarioResult(
        name=name,
        total_annual_cost=total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        annual_savings=annual_savings.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        savings_percent=savings_percent,
        spot_share_percent=spot_share_percent,
        risk_class=risk_class,
        risk_label=risk_label,
        breakdown=breakdown,
    )


def recommend(
    scenarios: List[ScenarioResult], risk_appetite: str
) -> Tuple[ScenarioResult, str]:
    """
    根据风险承受能力推荐方案。
    风险胃口：low / medium / high；默认 medium。
    """
    appetite = risk_appetite.lower()
    allowed = {
        "low": {"低"},
        "medium": {"低", "中"},
        "high": {"低", "中", "高"},
    }.get(appetite, {"低", "中"})

    ordered = sorted(scenarios, key=lambda s: s.total_annual_cost)
    for candidate in ordered:
        if candidate.risk_class in allowed:
            reason = f"成本最低且风险等级符合 {appetite} 胃口"
            return candidate, reason

    # 所有方案风险都超出胃口，返回成本最低者并提示
    return ordered[0], "所有方案风险均超出当前胃口，默认返回成本最低方案"


# ---------------------------------------------------------------------------
# 输出
# ---------------------------------------------------------------------------

def print_console(
    metadata: Dict[str, Any],
    scenarios: List[ScenarioResult],
    recommended: ScenarioResult,
    reason: str,
    currency: str,
) -> None:
    print("=" * 80)
    print("FinOps 承诺折扣优化报告")
    print("=" * 80)
    print(f"组织     : {metadata.get('organization', 'N/A')}")
    print(f"周期     : {metadata.get('period', 'N/A')}")
    print(f"币种     : {currency}")
    print(f"分析说明 : {metadata.get('description', '')}")
    print("-" * 80)

    print("\n场景对比")
    print(
        f"{'场景':<30} {'年度总成本':>14} {'年度节省':>14} "
        f"{'节省率':>10} {'Spot占比':>10} {'风险':>8}"
    )
    print("-" * 92)
    for s in scenarios:
        print(
            f"{s.name:<30} {fmt_currency(s.total_annual_cost):>14} "
            f"{fmt_currency(s.annual_savings):>14} {fmt_percent(s.savings_percent):>10} "
            f"{fmt_percent(s.spot_share_percent):>10} {s.risk_label:>8}"
        )

    print("\n" + "=" * 80)
    print("推荐方案")
    print("=" * 80)
    print(f"方案     : {recommended.name}")
    print(f"年度成本 : {fmt_currency(recommended.total_annual_cost)} {currency}")
    print(f"年度节省 : {fmt_currency(recommended.annual_savings)} {currency}")
    print(f"节省率   : {fmt_percent(recommended.savings_percent)}")
    print(f"风险等级 : {recommended.risk_label}")
    print(f"推荐理由 : {reason}")

    print("\n工作负载细分（推荐方案）")
    print(
        f"{'工作负载':<18} {'可中断':<8} {'分配类型':<20} "
        f"{'折扣率':>10} {'年度成本':>14}"
    )
    print("-" * 78)
    for row in recommended.breakdown:
        interrupt = "是" if row["can_interrupt"] else "否"
        print(
            f"{row['name']:<18} {interrupt:<8} {row['assigned_type']:<20} "
            f"{fmt_percent(row['discount_rate'] * 100):>10} "
            f"{fmt_currency(row['annual_cost']):>14}"
        )
    print("=" * 80)


def write_csv(
    path: Path,
    metadata: Dict[str, Any],
    scenarios: List[ScenarioResult],
    recommended: ScenarioResult,
    reason: str,
    currency: str,
) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)

        # 元数据
        writer.writerow(["FinOps Commitment Optimizer Report"])
        writer.writerow(["组织", metadata.get("organization", "N/A")])
        writer.writerow(["周期", metadata.get("period", "N/A")])
        writer.writerow(["币种", currency])
        writer.writerow(["描述", metadata.get("description", "")])
        writer.writerow([])

        # 场景汇总
        writer.writerow(
            ["场景", "年度总成本", "年度节省", "节省率(%)", "Spot成本占比(%)", "风险等级"]
        )
        for s in scenarios:
            writer.writerow(
                [
                    s.name,
                    fmt_currency(s.total_annual_cost),
                    fmt_currency(s.annual_savings),
                    fmt_currency(s.savings_percent),
                    fmt_currency(s.spot_share_percent),
                    s.risk_label,
                ]
            )
        writer.writerow([])

        # 推荐
        writer.writerow(["推荐方案", recommended.name])
        writer.writerow(["推荐年度总成本", fmt_currency(recommended.total_annual_cost)])
        writer.writerow(["推荐年度节省", fmt_currency(recommended.annual_savings)])
        writer.writerow(["推荐节省率(%)", fmt_currency(recommended.savings_percent)])
        writer.writerow(["推荐风险等级", recommended.risk_label])
        writer.writerow(["推荐理由", reason])
        writer.writerow([])

        # 推荐方案细分
        writer.writerow(
            ["工作负载ID", "工作负载", "可中断", "分配类型", "折扣率(%)", "年度成本"]
        )
        for row in recommended.breakdown:
            writer.writerow(
                [
                    row["workload_id"],
                    row["name"],
                    "是" if row["can_interrupt"] else "否",
                    row["assigned_type"],
                    fmt_currency(row["discount_rate"] * 100),
                    fmt_currency(row["annual_cost"]),
                ]
            )


def write_excel(
    path: Path,
    metadata: Dict[str, Any],
    scenarios: List[ScenarioResult],
    recommended: ScenarioResult,
    reason: str,
    currency: str,
) -> None:
    if not HAS_OPENPYXL:
        print(f"[警告] 未安装 openpyxl，跳过 Excel 输出：{path}", file=sys.stderr)
        return

    wb = openpyxl.Workbook()

    # 汇总 sheet
    ws_summary = wb.active
    ws_summary.title = "汇总"
    bold = Font(bold=True)
    ws_summary.append(["FinOps Commitment Optimizer Report"])
    ws_summary.append(["组织", metadata.get("organization", "N/A")])
    ws_summary.append(["周期", metadata.get("period", "N/A")])
    ws_summary.append(["币种", currency])
    ws_summary.append(["描述", metadata.get("description", "")])
    ws_summary.append([])
    ws_summary.append(["场景", "年度总成本", "年度节省", "节省率(%)", "Spot成本占比(%)", "风险等级"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = bold
    for s in scenarios:
        ws_summary.append(
            [
                s.name,
                float(s.total_annual_cost),
                float(s.annual_savings),
                float(s.savings_percent),
                float(s.spot_share_percent),
                s.risk_label,
            ]
        )
    ws_summary.append([])
    ws_summary.append(["推荐方案", recommended.name])
    ws_summary.append(["推荐年度总成本", float(recommended.total_annual_cost)])
    ws_summary.append(["推荐年度节省", float(recommended.annual_savings)])
    ws_summary.append(["推荐节省率(%)", float(recommended.savings_percent)])
    ws_summary.append(["推荐风险等级", recommended.risk_label])
    ws_summary.append(["推荐理由", reason])

    # 明细 sheet
    ws_detail = wb.create_sheet("推荐方案明细")
    ws_detail.append(
        ["工作负载ID", "工作负载", "可中断", "分配类型", "折扣率(%)", "年度成本"]
    )
    for cell in ws_detail[ws_detail.max_row]:
        cell.font = bold
    for row in recommended.breakdown:
        ws_detail.append(
            [
                row["workload_id"],
                row["name"],
                "是" if row["can_interrupt"] else "否",
                row["assigned_type"],
                float(row["discount_rate"] * 100),
                float(row["annual_cost"]),
            ]
        )

    # 各场景明细 sheet
    for s in scenarios:
        ws = wb.create_sheet(s.name.replace("/", "-").replace("（", "(").replace("）", ")")[:31])
        ws.append(["工作负载ID", "工作负载", "可中断", "分配类型", "折扣率(%)", "年度成本"])
        for cell in ws[ws.max_row]:
            cell.font = bold
        for row in s.breakdown:
            ws.append(
                [
                    row["workload_id"],
                    row["name"],
                    "是" if row["can_interrupt"] else "否",
                    row["assigned_type"],
                    float(row["discount_rate"] * 100),
                    float(row["annual_cost"]),
                ]
            )
        ws.append([])
        ws.append(
            ["年度总成本", float(s.total_annual_cost), "年度节省", float(s.annual_savings)]
        )

    wb.save(str(path))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="FinOps 承诺折扣优化器：比较全按需、RI/SP+Spot 最优组合、全 Spot 三种场景。"
    )
    parser.add_argument("--input", "-i", required=True, help="输入 YAML/JSON 文件路径")
    parser.add_argument("--csv", "-c", help="输出 CSV 文件路径")
    parser.add_argument("--excel", "-x", help="输出 Excel 文件路径（需安装 openpyxl）")
    parser.add_argument(
        "--risk-appetite",
        choices=["low", "medium", "high"],
        default=None,
        help="风险承受能力：low/medium/high（默认使用输入文件中的 assumptions.risk_appetite，否则 medium）",
    )
    return parser


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[错误] 输入文件不存在：{input_path}", file=sys.stderr)
        return 1

    raw = load_input(input_path)
    metadata = raw.get("metadata", {})
    currency = metadata.get("currency", "USD")
    risk_appetite = args.risk_appetite or raw.get("assumptions", {}).get("risk_appetite", "medium")

    discounts = parse_discounts(raw.get("discounts", {}))
    workloads = parse_workloads(raw.get("workloads", []))

    if not workloads:
        print("[错误] 输入中未定义任何工作负载（workloads）。", file=sys.stderr)
        return 1

    scenarios = build_scenarios(workloads, discounts)
    recommended, reason = recommend(scenarios, risk_appetite)

    print_console(metadata, scenarios, recommended, reason, currency)

    if args.csv:
        csv_path = Path(args.csv)
        write_csv(csv_path, metadata, scenarios, recommended, reason, currency)
        print(f"\n[CSV] 报告已写入：{csv_path.resolve()}")

    if args.excel:
        excel_path = Path(args.excel)
        write_excel(excel_path, metadata, scenarios, recommended, reason, currency)
        if HAS_OPENPYXL:
            print(f"[Excel] 报告已写入：{excel_path.resolve()}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
