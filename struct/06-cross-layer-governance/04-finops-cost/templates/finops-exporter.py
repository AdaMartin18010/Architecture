#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FinOps 四级成本分摊计算与导出器 (L1–L4 Allocation Exporter)

功能：
    1. 读取 YAML/JSON 格式的成本数据
    2. 按四级模型（资产级→项目级→组织级→生态级）执行分摊计算
    3. 导出为带公式的 Excel 文件（优先 openpyxl）或 CSV 文件

对齐来源：
    - FinOps Foundation Framework 2026
    - FinOps Foundation: Cost Allocation Capabilities
    - FOCUS (FinOps Open Cost and Usage Specification)

用法：
    python finops-exporter.py --input example-costs.yaml --output allocation.xlsx
    python finops-exporter.py --input example-costs.json --output allocation.csv

依赖：
    - PyYAML（YAML 输入时必需）
    - openpyxl（Excel 输出时优先尝试）
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# 可选依赖适配
# ---------------------------------------------------------------------------
try:
    import yaml

    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# 数据模型
# ---------------------------------------------------------------------------
@dataclass
class Asset:
    """L1 资产级复用组件/服务。"""

    id: str
    name: str
    layer: str
    description: str
    original_development_cost: Decimal
    monthly_maintenance: Decimal
    monthly_deployment: Decimal
    adaptation_adjustment_factor: Decimal
    usage_unit: str


@dataclass
class Project:
    """L2 项目级成本中心。"""

    id: str
    name: str
    description: str
    usage: Dict[str, Decimal]  # asset_id -> usage_quantity


@dataclass
class OrganizationalOverhead:
    """L3 组织级间接成本。"""

    coe_monthly_cost: Decimal
    platform_team_monthly_cost: Decimal
    shared_infrastructure_monthly: Decimal

    @property
    def total_monthly(self) -> Decimal:
        return self.coe_monthly_cost + self.platform_team_monthly_cost + self.shared_infrastructure_monthly


@dataclass
class RiskConfig:
    """L4 生态级风险成本配置。"""

    supply_chain_risk_score: Decimal
    security_investment_factor: Decimal
    license_compliance_reserve: Decimal
    vendor_lockin_reserve: Decimal

    @property
    def base_monthly_reserve(self) -> Decimal:
        return self.license_compliance_reserve + self.vendor_lockin_reserve

    def calculate_risk_cost(self, operational_base: Decimal) -> Decimal:
        """风险成本 = 准备金 + 运营成本 × (风险评分/10) × 安全投入系数。"""
        risk_multiplier = (self.supply_chain_risk_score / Decimal("10")) * self.security_investment_factor
        return self.base_monthly_reserve + (operational_base * risk_multiplier)


@dataclass
class AllocationResult:
    """完整的分摊计算结果。"""

    metadata: Dict[str, Any]
    # L1: asset_id -> monthly_direct_cost
    asset_direct_costs: Dict[str, Decimal]
    # L2: project_id -> {asset_id -> allocated_cost}
    project_asset_allocation: Dict[str, Dict[str, Decimal]]
    # L2: project_id -> total_direct_cost
    project_direct_costs: Dict[str, Decimal]
    # L3: project_id -> indirect_cost
    project_indirect_costs: Dict[str, Decimal]
    # L4: project_id -> risk_cost
    project_risk_costs: Dict[str, Decimal]
    # 汇总: project_id -> total_cost
    project_total_costs: Dict[str, Decimal]
    # 全局汇总
    total_direct: Decimal
    total_indirect: Decimal
    total_risk: Decimal
    grand_total: Decimal


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------
def load_data(path: Path) -> Dict[str, Any]:
    """从 YAML 或 JSON 文件加载成本数据。"""
    text = path.read_text(encoding="utf-8")
    suffix = path.suffix.lower()

    if suffix in (".yaml", ".yml"):
        if not HAS_YAML:
            raise RuntimeError(
                "读取 YAML 需要 PyYAML。请安装: pip install PyYAML，"
                "或将输入文件转换为 JSON 格式。"
            )
        return yaml.safe_load(text)
    elif suffix == ".json":
        return json.loads(text)
    else:
        # 尝试自动推断
        if text.strip().startswith(("{", "[")):
            return json.loads(text)
        if HAS_YAML:
            return yaml.safe_load(text)
        raise RuntimeError(f"无法识别文件格式: {suffix}。请使用 .yaml, .yml 或 .json。")


def parse_assets(raw: List[Dict[str, Any]]) -> List[Asset]:
    """解析资产列表。"""
    assets: List[Asset] = []
    for item in raw:
        assets.append(
            Asset(
                id=str(item["id"]),
                name=str(item["name"]),
                layer=str(item["layer"]),
                description=str(item.get("description", "")),
                original_development_cost=Decimal(str(item["original_development_cost"])),
                monthly_maintenance=Decimal(str(item["monthly_maintenance"])),
                monthly_deployment=Decimal(str(item["monthly_deployment"])),
                adaptation_adjustment_factor=Decimal(str(item["adaptation_adjustment_factor"])),
                usage_unit=str(item["usage_unit"]),
            )
        )
    return assets


def parse_projects(raw: List[Dict[str, Any]]) -> List[Project]:
    """解析项目列表。"""
    projects: List[Project] = []
    for item in raw:
        usage = {
            str(k): Decimal(str(v))
            for k, v in item.get("usage", {}).items()
        }
        projects.append(
            Project(
                id=str(item["id"]),
                name=str(item["name"]),
                description=str(item.get("description", "")),
                usage=usage,
            )
        )
    return projects


def parse_organizational(raw: Dict[str, Any]) -> OrganizationalOverhead:
    """解析组织级间接成本。"""
    return OrganizationalOverhead(
        coe_monthly_cost=Decimal(str(raw["coe_monthly_cost"])),
        platform_team_monthly_cost=Decimal(str(raw["platform_team_monthly_cost"])),
        shared_infrastructure_monthly=Decimal(str(raw["shared_infrastructure_monthly"])),
    )


def parse_risk(raw: Dict[str, Any]) -> RiskConfig:
    """解析风险配置。"""
    return RiskConfig(
        supply_chain_risk_score=Decimal(str(raw["supply_chain_risk_score"])),
        security_investment_factor=Decimal(str(raw["security_investment_factor"])),
        license_compliance_reserve=Decimal(str(raw["license_compliance_reserve"])),
        vendor_lockin_reserve=Decimal(str(raw["vendor_lockin_reserve"])),
    )


# ---------------------------------------------------------------------------
# 分摊计算引擎
# ---------------------------------------------------------------------------
def quantize(value: Decimal) -> Decimal:
    """将 Decimal 量化为两位小数（货币格式）。"""
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_allocation(
    assets: List[Asset],
    projects: List[Project],
    overhead: OrganizationalOverhead,
    risk: RiskConfig,
    metadata: Dict[str, Any],
) -> AllocationResult:
    """
    执行四级分摊计算。

    计算逻辑：
        L1 资产级:
            资产月度直接成本 = (原始开发成本 / 摊销月数) × AAF + 月维护 + 月部署

        L2 项目级:
            项目分摊 = Σ(资产月度直接成本 × 项目对该资产使用量 / 资产总使用量)

        L3 组织级:
            项目间接成本 = 总间接成本 × (项目复用资产数 / 所有项目复用资产数之和)

        L4 生态级:
            总风险成本 = 准备金 + (直接+间接)总成本 × (风险评分/10) × 安全投入系数
            项目风险成本 = 总风险成本 × (项目直接+间接成本 / 所有项目直接+间接成本之和)
    """
    amortization_months = Decimal(str(metadata.get("amortization_months", 36)))

    # ---- L1: 资产级直接成本 ------------------------------------------------
    asset_direct_costs: Dict[str, Decimal] = {}
    for asset in assets:
        amortized_dev = (asset.original_development_cost / amortization_months) * asset.adaptation_adjustment_factor
        monthly_direct = amortized_dev + asset.monthly_maintenance + asset.monthly_deployment
        asset_direct_costs[asset.id] = quantize(monthly_direct)

    # ---- L2: 项目级分摊 ----------------------------------------------------
    # 计算每个资产的总使用量
    asset_total_usage: Dict[str, Decimal] = {asset.id: Decimal("0") for asset in assets}
    for project in projects:
        for asset_id, qty in project.usage.items():
            if asset_id in asset_total_usage:
                asset_total_usage[asset_id] += qty

    project_asset_allocation: Dict[str, Dict[str, Decimal]] = {}
    project_direct_costs: Dict[str, Decimal] = {}

    for project in projects:
        allocation: Dict[str, Decimal] = {}
        total = Decimal("0")
        for asset in assets:
            asset_id = asset.id
            usage = project.usage.get(asset_id, Decimal("0"))
            total_usage = asset_total_usage[asset_id]
            if total_usage > 0:
                cost = asset_direct_costs[asset_id] * usage / total_usage
            else:
                cost = Decimal("0")
            allocation[asset_id] = quantize(cost)
            total += cost
        project_asset_allocation[project.id] = allocation
        project_direct_costs[project.id] = quantize(total)

    total_direct = sum(project_direct_costs.values(), Decimal("0"))

    # ---- L3: 组织级间接成本 ------------------------------------------------
    total_overhead = overhead.total_monthly
    # 计算每个项目复用资产数
    project_asset_counts: Dict[str, int] = {}
    for project in projects:
        project_asset_counts[project.id] = len(project.usage)
    total_asset_usage_count = sum(project_asset_counts.values())

    project_indirect_costs: Dict[str, Decimal] = {}
    if total_asset_usage_count > 0:
        for project in projects:
            ratio = Decimal(str(project_asset_counts[project.id])) / Decimal(str(total_asset_usage_count))
            project_indirect_costs[project.id] = quantize(total_overhead * ratio)
    else:
        for project in projects:
            project_indirect_costs[project.id] = Decimal("0")

    total_indirect = sum(project_indirect_costs.values(), Decimal("0"))

    # ---- L4: 生态级风险成本 ------------------------------------------------
    operational_base = total_direct + total_indirect
    total_risk = quantize(risk.calculate_risk_cost(operational_base))

    project_risk_costs: Dict[str, Decimal] = {}
    if operational_base > 0:
        for project in projects:
            project_operational = project_direct_costs[project.id] + project_indirect_costs[project.id]
            ratio = project_operational / operational_base
            project_risk_costs[project.id] = quantize(total_risk * ratio)
    else:
        for project in projects:
            project_risk_costs[project.id] = Decimal("0")

    # ---- 汇总 --------------------------------------------------------------
    project_total_costs: Dict[str, Decimal] = {}
    for project in projects:
        total = (
            project_direct_costs[project.id]
            + project_indirect_costs[project.id]
            + project_risk_costs[project.id]
        )
        project_total_costs[project.id] = quantize(total)

    grand_total = sum(project_total_costs.values(), Decimal("0"))

    return AllocationResult(
        metadata=metadata,
        asset_direct_costs=asset_direct_costs,
        project_asset_allocation=project_asset_allocation,
        project_direct_costs=project_direct_costs,
        project_indirect_costs=project_indirect_costs,
        project_risk_costs=project_risk_costs,
        project_total_costs=project_total_costs,
        total_direct=quantize(total_direct),
        total_indirect=quantize(total_indirect),
        total_risk=quantize(total_risk),
        grand_total=quantize(grand_total),
    )


# ---------------------------------------------------------------------------
# 报表输出: Excel (openpyxl)
# ---------------------------------------------------------------------------
def write_excel(
    path: Path,
    assets: List[Asset],
    projects: List[Project],
    overhead: OrganizationalOverhead,
    risk: RiskConfig,
    result: AllocationResult,
) -> None:
    """使用 openpyxl 导出带公式的 Excel 文件。"""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl 未安装")

    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    num_fmt = '#,##0.00'
    pct_fmt = '0.00%'

    def style_header(ws, row: int, cols: int):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    if val_len > max_length:
                        max_length = val_len
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    # ---- Sheet 1: 摘要 -----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "摘要"
    summary_data = [
        ["指标", "值 (USD)"],
        ["分摊周期", result.metadata.get("period", "")],
        ["组织", result.metadata.get("organization", "")],
        ["摊销月数", str(result.metadata.get("amortization_months", 36))],
        ["总直接成本", float(result.total_direct)],
        ["总间接成本", float(result.total_indirect)],
        ["总风险成本", float(result.total_risk)],
        ["成本总计", float(result.grand_total)],
        ["", ""],
        ["组织级间接成本明细", ""],
        ["CoE 月度成本", float(overhead.coe_monthly_cost)],
        ["平台团队月度成本", float(overhead.platform_team_monthly_cost)],
        ["共享基础设施月度成本", float(overhead.shared_infrastructure_monthly)],
        ["", ""],
        ["风险配置", ""],
        ["供应链风险评分 (1-10)", float(risk.supply_chain_risk_score)],
        ["安全投入系数", float(risk.security_investment_factor)],
        ["许可证合规准备金", float(risk.license_compliance_reserve)],
        ["供应商锁定准备金", float(risk.vendor_lockin_reserve)],
    ]
    for r_idx, row in enumerate(summary_data, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws_summary.cell(row=r_idx, column=c_idx, value=val)
            if isinstance(val, float):
                ws_summary.cell(row=r_idx, column=c_idx).number_format = num_fmt
    style_header(ws_summary, 1, 2)
    auto_width(ws_summary)

    # ---- Sheet 2: 资产级成本 (L1) ------------------------------------------
    ws_assets = wb.create_sheet("L1-资产级成本")
    asset_headers = [
        "资产ID", "资产名称", "层级", "原始开发成本", "AAF",
        "月度维护", "月度部署", "月度直接成本", "等效改编成本", "单位"
    ]
    ws_assets.append(asset_headers)
    style_header(ws_assets, 1, len(asset_headers))

    for asset in assets:
        monthly_direct = result.asset_direct_costs[asset.id]
        adaptation_cost = quantize(
            asset.original_development_cost * asset.adaptation_adjustment_factor
            / Decimal(str(result.metadata.get("amortization_months", 36)))
        )
        ws_assets.append([
            asset.id,
            asset.name,
            asset.layer,
            float(asset.original_development_cost),
            float(asset.adaptation_adjustment_factor),
            float(asset.monthly_maintenance),
            float(asset.monthly_deployment),
            float(monthly_direct),
            float(adaptation_cost),
            asset.usage_unit,
        ])
    auto_width(ws_assets)

    # ---- Sheet 3: 项目级分摊 (L2) ------------------------------------------
    ws_proj = wb.create_sheet("L2-项目级分摊")
    proj_headers = ["项目ID", "项目名称"] + [a.name for a in assets] + ["直接成本合计"]
    ws_proj.append(proj_headers)
    style_header(ws_proj, 1, len(proj_headers))

    for project in projects:
        row = [project.id, project.name]
        for asset in assets:
            row.append(float(result.project_asset_allocation[project.id].get(asset.id, Decimal("0"))))
        # 合计列使用 SUM 公式
        start_col = 3
        end_col = start_col + len(assets) - 1
        row_data = row + [None]  # 占位，后面用公式替换
        ws_proj.append(row_data[:-1] + [f"=SUM({get_column_letter(start_col)}{ws_proj.max_row}:{get_column_letter(end_col)}{ws_proj.max_row})"])

    # 添加汇总行（带公式）
    total_row_idx = ws_proj.max_row + 1
    ws_proj.cell(row=total_row_idx, column=1, value="合计")
    ws_proj.cell(row=total_row_idx, column=2, value="")
    for col_idx in range(3, len(proj_headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws_proj.cell(row=total_row_idx, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})")
    auto_width(ws_proj)

    # ---- Sheet 4: 组织级间接成本 (L3) --------------------------------------
    ws_indirect = wb.create_sheet("L3-组织级间接成本")
    ws_indirect.append(["项目ID", "项目名称", "复用资产数", "占比", "间接成本 (USD)"])
    style_header(ws_indirect, 1, 5)

    total_asset_usage_count = sum(len(p.usage) for p in projects)
    for project in projects:
        asset_count = len(project.usage)
        ratio = asset_count / total_asset_usage_count if total_asset_usage_count else 0
        ws_indirect.append([
            project.id,
            project.name,
            asset_count,
            ratio,
            float(result.project_indirect_costs[project.id]),
        ])
    auto_width(ws_indirect)

    # ---- Sheet 5: 风险成本 (L4) --------------------------------------------
    ws_risk = wb.create_sheet("L4-风险成本")
    ws_risk.append(["项目ID", "项目名称", "直接+间接成本", "占比", "风险成本 (USD)"])
    style_header(ws_risk, 1, 5)

    operational_base = result.total_direct + result.total_indirect
    for project in projects:
        proj_op = result.project_direct_costs[project.id] + result.project_indirect_costs[project.id]
        ratio = float(proj_op / operational_base) if operational_base > 0 else 0.0
        ws_risk.append([
            project.id,
            project.name,
            float(proj_op),
            ratio,
            float(result.project_risk_costs[project.id]),
        ])
    auto_width(ws_risk)

    # ---- Sheet 6: 最终报告 -------------------------------------------------
    ws_final = wb.create_sheet("最终报告")
    ws_final.append(["项目ID", "项目名称", "直接成本", "间接成本", "风险成本", "总成本", "占总成本比例"])
    style_header(ws_final, 1, 7)

    for project in projects:
        direct = result.project_direct_costs[project.id]
        indirect = result.project_indirect_costs[project.id]
        risk_c = result.project_risk_costs[project.id]
        total = result.project_total_costs[project.id]
        ratio = float(total / result.grand_total) if result.grand_total > 0 else 0.0
        ws_final.append([
            project.id,
            project.name,
            float(direct),
            float(indirect),
            float(risk_c),
            float(total),
            ratio,
        ])

    # 汇总行
    final_total_row = ws_final.max_row + 1
    ws_final.cell(row=final_total_row, column=1, value="合计")
    for col_idx in range(3, 7):
        col_letter = get_column_letter(col_idx)
        ws_final.cell(row=final_total_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{final_total_row - 1})")
    ws_final.cell(row=final_total_row, column=7, value=1.0)
    auto_width(ws_final)

    wb.save(str(path))


# ---------------------------------------------------------------------------
# 报表输出: CSV (fallback)
# ---------------------------------------------------------------------------
def write_csv(
    path: Path,
    assets: List[Asset],
    projects: List[Project],
    overhead: OrganizationalOverhead,
    risk: RiskConfig,
    result: AllocationResult,
) -> None:
    """
    当 openpyxl 不可用时，生成多个 CSV 文件。

    如果输出路径以 .xlsx 结尾，则自动创建一个同名目录存放 CSV 文件。
    否则按给定路径名创建目录。
    """
    if path.suffix.lower() == ".xlsx":
        dir_path = path.with_suffix("")
    else:
        dir_path = path

    dir_path.mkdir(parents=True, exist_ok=True)

    # CSV 1: 摘要
    with open(dir_path / "01_summary.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["指标", "值"])
        writer.writerow(["分摊周期", result.metadata.get("period", "")])
        writer.writerow(["组织", result.metadata.get("organization", "")])
        writer.writerow(["摊销月数", str(result.metadata.get("amortization_months", 36))])
        writer.writerow(["总直接成本", str(result.total_direct)])
        writer.writerow(["总间接成本", str(result.total_indirect)])
        writer.writerow(["总风险成本", str(result.total_risk)])
        writer.writerow(["成本总计", str(result.grand_total)])

    # CSV 2: 资产级成本
    with open(dir_path / "02_l1_assets.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "资产ID", "资产名称", "层级", "原始开发成本", "AAF",
            "月度维护", "月度部署", "月度直接成本", "单位"
        ])
        for asset in assets:
            writer.writerow([
                asset.id,
                asset.name,
                asset.layer,
                str(asset.original_development_cost),
                str(asset.adaptation_adjustment_factor),
                str(asset.monthly_maintenance),
                str(asset.monthly_deployment),
                str(result.asset_direct_costs[asset.id]),
                asset.usage_unit,
            ])

    # CSV 3: 项目级分摊
    with open(dir_path / "03_l2_project_allocation.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        headers = ["项目ID", "项目名称"] + [a.name for a in assets] + ["直接成本合计"]
        writer.writerow(headers)
        for project in projects:
            row = [project.id, project.name]
            total = Decimal("0")
            for asset in assets:
                val = result.project_asset_allocation[project.id].get(asset.id, Decimal("0"))
                row.append(str(val))
                total += val
            row.append(str(quantize(total)))
            writer.writerow(row)

    # CSV 4: 组织级间接成本
    with open(dir_path / "04_l3_indirect.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["项目ID", "项目名称", "复用资产数", "占比", "间接成本"])
        total_asset_usage_count = sum(len(p.usage) for p in projects)
        for project in projects:
            asset_count = len(project.usage)
            ratio = Decimal(str(asset_count)) / Decimal(str(total_asset_usage_count)) if total_asset_usage_count else Decimal("0")
            writer.writerow([
                project.id,
                project.name,
                asset_count,
                str(ratio.quantize(Decimal("0.0001"))),
                str(result.project_indirect_costs[project.id]),
            ])

    # CSV 5: 风险成本
    with open(dir_path / "05_l4_risk.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["项目ID", "项目名称", "直接+间接成本", "占比", "风险成本"])
        operational_base = result.total_direct + result.total_indirect
        for project in projects:
            proj_op = result.project_direct_costs[project.id] + result.project_indirect_costs[project.id]
            ratio = proj_op / operational_base if operational_base > 0 else Decimal("0")
            writer.writerow([
                project.id,
                project.name,
                str(proj_op),
                str(ratio.quantize(Decimal("0.0001"))),
                str(result.project_risk_costs[project.id]),
            ])

    # CSV 6: 最终报告
    with open(dir_path / "06_final_report.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["项目ID", "项目名称", "直接成本", "间接成本", "风险成本", "总成本", "占总成本比例"])
        for project in projects:
            direct = result.project_direct_costs[project.id]
            indirect = result.project_indirect_costs[project.id]
            risk_c = result.project_risk_costs[project.id]
            total = result.project_total_costs[project.id]
            ratio = total / result.grand_total if result.grand_total > 0 else Decimal("0")
            writer.writerow([
                project.id,
                project.name,
                str(direct),
                str(indirect),
                str(risk_c),
                str(total),
                str(ratio.quantize(Decimal("0.0001"))),
            ])
        writer.writerow([
            "合计", "",
            str(result.total_direct),
            str(result.total_indirect),
            str(result.total_risk),
            str(result.grand_total),
            "1.0000",
        ])


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description="FinOps 四级成本分摊计算与导出器 (L1–L4 Allocation Exporter)"
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        type=Path,
        help="输入 YAML/JSON 文件路径",
    )
    parser.add_argument(
        "--output", "-o",
        required=True,
        type=Path,
        help="输出文件路径 (.xlsx 优先，若 openpyxl 不可用则生成 CSV 目录)",
    )
    parser.add_argument(
        "--format",
        choices=["auto", "excel", "csv"],
        default="auto",
        help="输出格式: auto(默认优先Excel), excel(强制), csv(强制)",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"错误: 输入文件不存在: {args.input}", file=sys.stderr)
        return 1

    try:
        raw_data = load_data(args.input)
    except Exception as exc:
        print(f"错误: 无法加载输入文件: {exc}", file=sys.stderr)
        return 1

    # 解析数据
    try:
        metadata = raw_data.get("metadata", {})
        assets = parse_assets(raw_data.get("assets", []))
        projects = parse_projects(raw_data.get("projects", []))
        overhead = parse_organizational(raw_data.get("organizational_overhead", {}))
        risk = parse_risk(raw_data.get("risk", {}))
    except KeyError as exc:
        print(f"错误: 输入文件缺少必要字段: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"错误: 数据解析失败: {exc}", file=sys.stderr)
        return 1

    if not assets:
        print("错误: 未找到资产数据", file=sys.stderr)
        return 1
    if not projects:
        print("错误: 未找到项目数据", file=sys.stderr)
        return 1

    # 执行计算
    try:
        result = calculate_allocation(assets, projects, overhead, risk, metadata)
    except Exception as exc:
        print(f"错误: 分摊计算失败: {exc}", file=sys.stderr)
        return 1

    # 决定输出格式
    use_excel = False
    if args.format == "excel":
        if not HAS_OPENPYXL:
            print("错误: 强制指定 Excel 输出但 openpyxl 未安装。请运行: pip install openpyxl", file=sys.stderr)
            return 1
        use_excel = True
    elif args.format == "csv":
        use_excel = False
    else:  # auto
        use_excel = HAS_OPENPYXL

    # 导出
    try:
        if use_excel:
            write_excel(args.output, assets, projects, overhead, risk, result)
            print(f"Excel 报告已生成: {args.output}")
        else:
            write_csv(args.output, assets, projects, overhead, risk, result)
            if args.output.suffix.lower() == ".xlsx":
                csv_dir = args.output.with_suffix("")
            else:
                csv_dir = args.output
            print(f"CSV 报告已生成至目录: {csv_dir}/")
            print("提示: 这些 CSV 文件可直接用 Excel / WPS 打开。")
    except Exception as exc:
        print(f"错误: 导出失败: {exc}", file=sys.stderr)
        return 1

    # 控制台摘要
    print("\n===== FinOps 四级分摊摘要 =====")
    print(f"周期: {metadata.get('period', 'N/A')} | 组织: {metadata.get('organization', 'N/A')}")
    print(f"总直接成本: {result.total_direct} | 总间接成本: {result.total_indirect} | 总风险成本: {result.total_risk}")
    print(f"成本总计: {result.grand_total}")
    print("\n按项目汇总:")
    print(f"{'项目':<20} {'直接':>12} {'间接':>12} {'风险':>12} {'总计':>12}")
    print("-" * 72)
    for project in projects:
        print(
            f"{project.name:<20} "
            f"{result.project_direct_costs[project.id]:>12} "
            f"{result.project_indirect_costs[project.id]:>12} "
            f"{result.project_risk_costs[project.id]:>12} "
            f"{result.project_total_costs[project.id]:>12}"
        )

    return 0


if __name__ == "__main__":
    sys.exit(main())
