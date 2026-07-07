#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI GPU 成本分摊计算器 (AI GPU Cost Allocation Calculator)

功能：
    1. 读取 YAML/JSON 输入，描述 AI 工作负载（GPU、Token、存储、网络、日志）。
    2. 计算 GPU 总成本、Token 成本（input/output/总量）、附加成本。
    3. 按团队/项目/模型分摊共享 GPU 与平台服务成本。
    4. 输出每千次推理成本、每百万 token 成本、每 GPU 小时成本等单位经济学指标。
    5. 支持控制台、CSV、可选 Excel 三种输出格式。

用法：
    python ai-gpu-cost-calculator.py --input example-ai-gpu-cost.yaml
    python ai-gpu-cost-calculator.py --input example-ai-gpu-cost.yaml \
        --output report --format all

命令行参数：
    --input   输入文件路径（YAML 或 JSON）
    --output  输出文件基础路径（不含扩展名）
    --format  输出格式：csv | excel | all（默认 all；控制台始终输出）

依赖策略（标准库优先）：
    - YAML 解析优先使用 PyYAML；未安装时仅支持 JSON 输入。
    - Excel 优先使用 openpyxl；未安装时尝试 xlsxwriter；两者皆无时降级为仅 CSV/控制台。

对齐来源：
    - FinOps Foundation: AI Cost Management / Token Economics / Cost Allocation
    - GSF SCI for AI（Software Carbon Intensity for AI）
    - 本项目 ai-cost-allocation.md
    - 本项目 unit-economics.md
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import yaml  # type: ignore
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlsxwriter  # type: ignore
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False


# ---------------------------------------------------------------------------
# 数据模型
# ---------------------------------------------------------------------------

@dataclass
class Workload:
    id: str
    name: str
    team: str
    project: str
    model_id: str
    workload_type: str
    gpu_hours: Decimal
    memory_gb_hours: Decimal
    requests: Decimal
    input_tokens: Decimal
    output_tokens: Decimal
    storage_gb: Decimal
    network_gb: Decimal
    logs_gb: Decimal
    allocation_weight: Decimal

    @property
    def total_tokens(self) -> Decimal:
        return self.input_tokens + self.output_tokens


@dataclass
class AllocationResult:
    workload: Workload
    gpu_allocated_cost: Decimal
    input_token_cost: Decimal
    output_token_cost: Decimal
    storage_cost: Decimal
    network_cost: Decimal
    logs_cost: Decimal
    platform_allocated_cost: Decimal

    @property
    def total_token_cost(self) -> Decimal:
        return self.input_token_cost + self.output_token_cost

    @property
    def extra_cost(self) -> Decimal:
        return self.storage_cost + self.network_cost + self.logs_cost

    @property
    def total_cost(self) -> Decimal:
        return (
            self.gpu_allocated_cost
            + self.total_token_cost
            + self.extra_cost
            + self.platform_allocated_cost
        )


# ---------------------------------------------------------------------------
# 数值辅助
# ---------------------------------------------------------------------------

def d(value: Any) -> Decimal:
    """将常见数值类型安全转换为 Decimal。"""
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def fmt_currency(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def fmt_number(value: Optional[Decimal]) -> str:
    if value is None:
        return "N/A"
    return str(value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))


def safe_div(numerator: Decimal, denominator: Decimal) -> Optional[Decimal]:
    if denominator == 0:
        return None
    return numerator / denominator


# ---------------------------------------------------------------------------
# 输入解析
# ---------------------------------------------------------------------------

def load_input(path: Path) -> Dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    suffix = path.suffix.lower()

    if suffix in {".yaml", ".yml"}:
        if not HAS_YAML:
            print(
                "错误：检测到 YAML 输入但未安装 PyYAML。"
                "请运行 'pip install pyyaml'，或改用 JSON 输入。",
                file=sys.stderr,
            )
            sys.exit(2)
        return yaml.safe_load(text) or {}

    if suffix == ".json":
        return json.loads(text)

    # 未识别扩展名时自动探测
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        if HAS_YAML:
            return yaml.safe_load(text) or {}
        print("错误：无法解析输入文件，请提供 YAML/JSON 格式。", file=sys.stderr)
        sys.exit(2)


def parse_workloads(raw: Dict[str, Any]) -> List[Workload]:
    workloads: List[Workload] = []
    for item in raw.get("workloads", []):
        tokens = item.get("tokens", {})
        extra = item.get("extra_usage", {})
        workloads.append(
            Workload(
                id=str(item.get("id", "")),
                name=str(item.get("name", "")),
                team=str(item.get("team", "")),
                project=str(item.get("project", "")),
                model_id=str(item.get("model_id", "")),
                workload_type=str(item.get("workload_type", "")),
                gpu_hours=d(item.get("gpu_hours", 0)),
                memory_gb_hours=d(item.get("memory_gb_hours", 0)),
                requests=d(item.get("requests", 0)),
                input_tokens=d(tokens.get("input", 0)),
                output_tokens=d(tokens.get("output", 0)),
                storage_gb=d(extra.get("storage_gb", 0)),
                network_gb=d(extra.get("network_gb", 0)),
                logs_gb=d(extra.get("logs_gb", 0)),
                allocation_weight=d(item.get("allocation_weight", 1)),
            )
        )
    return workloads


# ---------------------------------------------------------------------------
# 计算引擎
# ---------------------------------------------------------------------------

def calculate_gpu_cluster_cost(raw: Dict[str, Any]) -> Decimal:
    cluster = raw.get("gpu_cluster", {})
    gpu_count = d(cluster.get("gpu_count", 0))
    hourly_price = d(cluster.get("hourly_price_per_gpu", 0))
    hours = d(cluster.get("hours", 0))
    overhead = d(cluster.get("scheduler_overhead_ratio", 0))
    base_cost = gpu_count * hourly_price * hours
    return base_cost * (Decimal("1") + overhead)


def allocation_driver_value(wl: Workload, method: str) -> Decimal:
    if method == "gpu_hours":
        return wl.gpu_hours
    if method == "memory_gb_hours":
        return wl.memory_gb_hours
    if method == "tokens":
        return wl.total_tokens
    if method == "requests":
        return wl.requests
    if method == "manual":
        return wl.allocation_weight
    return wl.gpu_hours


def allocate_cost(
    pool_cost: Decimal,
    workloads: List[Workload],
    method: str,
) -> List[Decimal]:
    drivers = [allocation_driver_value(wl, method) for wl in workloads]
    total_driver = sum(drivers)
    if total_driver == 0:
        return [Decimal("0") for _ in workloads]
    return [
        (pool_cost * drv / total_driver).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        for drv in drivers
    ]


def calculate(raw: Dict[str, Any]) -> Tuple[Decimal, List[AllocationResult], Dict[str, Decimal]]:
    workloads = parse_workloads(raw)
    if not workloads:
        print("警告：输入中未包含任何 workload。", file=sys.stderr)

    cluster = raw.get("gpu_cluster", {})
    allocation_method = str(cluster.get("allocation_method", "gpu_hours"))
    gpu_total_cost = calculate_gpu_cluster_cost(raw)

    gpu_allocations = allocate_cost(gpu_total_cost, workloads, allocation_method)

    llm = raw.get("llm_pricing", {})
    input_price = d(llm.get("input_token_price_per_1k", 0))
    output_price = d(llm.get("output_token_price_per_1k", 0))

    extra = raw.get("extra_costs", {})
    storage_price = d(extra.get("storage", {}).get("unit_price", 0))
    network_price = d(extra.get("network", {}).get("unit_price", 0))
    logs_price = d(extra.get("logs", {}).get("unit_price", 0))

    platform = raw.get("platform_services", {})
    platform_cost = d(platform.get("total_monthly_cost", 0))
    platform_driver = str(platform.get("allocation_driver", "tokens"))
    platform_allocations = allocate_cost(platform_cost, workloads, platform_driver)

    results: List[AllocationResult] = []
    for wl, gpu_alloc, platform_alloc in zip(
        workloads, gpu_allocations, platform_allocations
    ):
        results.append(
            AllocationResult(
                workload=wl,
                gpu_allocated_cost=gpu_alloc,
                input_token_cost=(wl.input_tokens / Decimal("1000")) * input_price,
                output_token_cost=(wl.output_tokens / Decimal("1000")) * output_price,
                storage_cost=wl.storage_gb * storage_price,
                network_cost=wl.network_gb * network_price,
                logs_cost=wl.logs_gb * logs_price,
                platform_allocated_cost=platform_alloc,
            )
        )

    summary: Dict[str, Decimal] = {
        "gpu_total_cost": gpu_total_cost,
        "total_input_token_cost": sum(r.input_token_cost for r in results),
        "total_output_token_cost": sum(r.output_token_cost for r in results),
        "total_token_cost": sum(r.total_token_cost for r in results),
        "total_storage_cost": sum(r.storage_cost for r in results),
        "total_network_cost": sum(r.network_cost for r in results),
        "total_logs_cost": sum(r.logs_cost for r in results),
        "total_extra_cost": sum(r.extra_cost for r in results),
        "total_platform_cost": platform_cost,
        "total_cost": sum(r.total_cost for r in results),
        "total_gpu_hours": sum(wl.gpu_hours for wl in workloads),
        "total_requests": sum(wl.requests for wl in workloads),
        "total_input_tokens": sum(wl.input_tokens for wl in workloads),
        "total_output_tokens": sum(wl.output_tokens for wl in workloads),
        "total_tokens": sum(wl.total_tokens for wl in workloads),
    }

    return gpu_total_cost, results, summary


# ---------------------------------------------------------------------------
# 控制台输出
# ---------------------------------------------------------------------------

def print_report(
    raw: Dict[str, Any],
    gpu_total_cost: Decimal,
    results: List[AllocationResult],
    summary: Dict[str, Decimal],
) -> None:
    meta = raw.get("metadata", {})
    cluster = raw.get("gpu_cluster", {})

    print("=" * 80)
    print("AI GPU 成本分摊报告")
    print("=" * 80)
    print(f"计算周期 : {meta.get('period', 'N/A')}")
    print(f"货币     : {meta.get('currency', 'USD')}")
    print(f"负责人   : {meta.get('owner', 'N/A')}")
    print(f"描述     : {meta.get('description', '')}")
    print(f"GPU 类型 : {cluster.get('gpu_type', 'N/A')} x {cluster.get('gpu_count', 0)}")
    print(f"分摊键   : {cluster.get('allocation_method', 'gpu_hours')}")
    print()

    print("-" * 80)
    print("汇总成本")
    print("-" * 80)
    print(f"  GPU 集群总成本           : {fmt_currency(summary['gpu_total_cost'])}")
    print(f"  Token 输入成本           : {fmt_currency(summary['total_input_token_cost'])}")
    print(f"  Token 输出成本           : {fmt_currency(summary['total_output_token_cost'])}")
    print(f"  Token 总成本             : {fmt_currency(summary['total_token_cost'])}")
    print(f"  存储成本                 : {fmt_currency(summary['total_storage_cost'])}")
    print(f"  网络成本                 : {fmt_currency(summary['total_network_cost'])}")
    print(f"  日志成本                 : {fmt_currency(summary['total_logs_cost'])}")
    print(f"  平台服务分摊成本         : {fmt_currency(summary['total_platform_cost'])}")
    print(f"  总成本                   : {fmt_currency(summary['total_cost'])}")
    print()

    print("-" * 80)
    print("单位经济学指标")
    print("-" * 80)
    cost_per_1k_requests = safe_div(
        summary["total_cost"], summary["total_requests"] / Decimal("1000")
    )
    cost_per_1m_tokens = safe_div(
        summary["total_cost"], summary["total_tokens"] / Decimal("1000000")
    )
    cost_per_gpu_hour = safe_div(
        summary["gpu_total_cost"], summary["total_gpu_hours"]
    )
    cost_per_1k_input = safe_div(
        summary["total_input_token_cost"],
        summary["total_input_tokens"] / Decimal("1000"),
    )
    cost_per_1k_output = safe_div(
        summary["total_output_token_cost"],
        summary["total_output_tokens"] / Decimal("1000"),
    )
    cost_per_1k_total_tokens = safe_div(
        summary["total_token_cost"],
        summary["total_tokens"] / Decimal("1000"),
    )

    print(f"  每千次推理成本           : {fmt_number(cost_per_1k_requests)}")
    print(f"  每百万 token 成本        : {fmt_number(cost_per_1m_tokens)}")
    print(f"  每 GPU 小时成本          : {fmt_number(cost_per_gpu_hour)}")
    print(f"  每千输入 token 成本      : {fmt_number(cost_per_1k_input)}")
    print(f"  每千输出 token 成本      : {fmt_number(cost_per_1k_output)}")
    print(f"  每千总 token 成本        : {fmt_number(cost_per_1k_total_tokens)}")
    print()

    print("-" * 80)
    print("Workload 明细")
    print("-" * 80)
    header = (
        f"{'ID':<12} {'Workload':<22} {'Team':<26} {'GPU':>10} "
        f"{'Token':>10} {'Extra':>10} {'Platform':>10} {'Total':>12}"
    )
    print(header)
    print("-" * len(header))
    for r in results:
        print(
            f"{r.workload.id:<12} {r.workload.name:<22} {r.workload.team:<26} "
            f"{fmt_currency(r.gpu_allocated_cost):>10} "
            f"{fmt_currency(r.total_token_cost):>10} "
            f"{fmt_currency(r.extra_cost):>10} "
            f"{fmt_currency(r.platform_allocated_cost):>10} "
            f"{fmt_currency(r.total_cost):>12}"
        )
    print()

    print_aggregation("按团队分摊", results, lambda r: r.workload.team)
    print_aggregation("按项目分摊", results, lambda r: r.workload.project)
    print_aggregation("按模型分摊", results, lambda r: r.workload.model_id)


def print_aggregation(
    title: str,
    results: List[AllocationResult],
    key_func,
) -> None:
    groups: Dict[str, Decimal] = {}
    for r in results:
        key = key_func(r)
        groups[key] = groups.get(key, Decimal("0")) + r.total_cost

    print("-" * 80)
    print(title)
    print("-" * 80)
    for key, cost in sorted(groups.items(), key=lambda x: -x[1]):
        share = safe_div(cost, sum(groups.values()))
        share_str = f"{share:.2%}" if share is not None else "N/A"
        print(f"  {key:<30} {fmt_currency(cost):>12}  ({share_str})")
    print()


# ---------------------------------------------------------------------------
# CSV 输出
# ---------------------------------------------------------------------------

def write_csv_report(
    output_base: Path,
    results: List[AllocationResult],
    summary: Dict[str, Decimal],
) -> Path:
    csv_path = Path(str(output_base) + ".csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        # 汇总
        writer.writerow(["section", "metric", "value"])
        writer.writerow(["summary", "gpu_total_cost", fmt_currency(summary["gpu_total_cost"])])
        writer.writerow(["summary", "total_input_token_cost", fmt_currency(summary["total_input_token_cost"])])
        writer.writerow(["summary", "total_output_token_cost", fmt_currency(summary["total_output_token_cost"])])
        writer.writerow(["summary", "total_token_cost", fmt_currency(summary["total_token_cost"])])
        writer.writerow(["summary", "total_extra_cost", fmt_currency(summary["total_extra_cost"])])
        writer.writerow(["summary", "total_platform_cost", fmt_currency(summary["total_platform_cost"])])
        writer.writerow(["summary", "total_cost", fmt_currency(summary["total_cost"])])
        writer.writerow(["summary", "total_gpu_hours", str(summary["total_gpu_hours"])])
        writer.writerow(["summary", "total_requests", str(summary["total_requests"])])
        writer.writerow(["summary", "total_tokens", str(summary["total_tokens"])])
        writer.writerow([])

        # Workload 明细
        headers = [
            "workload_id", "workload_name", "team", "project", "model_id",
            "workload_type", "gpu_hours", "memory_gb_hours", "requests",
            "input_tokens", "output_tokens", "total_tokens",
            "gpu_allocated_cost", "input_token_cost", "output_token_cost",
            "total_token_cost", "storage_cost", "network_cost", "logs_cost",
            "extra_cost", "platform_allocated_cost", "total_cost",
        ]
        writer.writerow(["workload"] + headers)
        for r in results:
            writer.writerow([
                "workload",
                r.workload.id,
                r.workload.name,
                r.workload.team,
                r.workload.project,
                r.workload.model_id,
                r.workload.workload_type,
                str(r.workload.gpu_hours),
                str(r.workload.memory_gb_hours),
                str(r.workload.requests),
                str(r.workload.input_tokens),
                str(r.workload.output_tokens),
                str(r.workload.total_tokens),
                fmt_currency(r.gpu_allocated_cost),
                fmt_currency(r.input_token_cost),
                fmt_currency(r.output_token_cost),
                fmt_currency(r.total_token_cost),
                fmt_currency(r.storage_cost),
                fmt_currency(r.network_cost),
                fmt_currency(r.logs_cost),
                fmt_currency(r.extra_cost),
                fmt_currency(r.platform_allocated_cost),
                fmt_currency(r.total_cost),
            ])
        writer.writerow([])

        # 分摊聚合
        writer.writerow(["aggregation", "dimension", "key", "cost"])
        for title, key_func in [
            ("team", lambda r: r.workload.team),
            ("project", lambda r: r.workload.project),
            ("model", lambda r: r.workload.model_id),
        ]:
            groups: Dict[str, Decimal] = {}
            for r in results:
                key = key_func(r)
                groups[key] = groups.get(key, Decimal("0")) + r.total_cost
            for key, cost in sorted(groups.items(), key=lambda x: -x[1]):
                writer.writerow(["aggregation", title, key, fmt_currency(cost)])

    return csv_path


# ---------------------------------------------------------------------------
# Excel 输出
# ---------------------------------------------------------------------------

def _write_excel_with_openpyxl(
    xlsx_path: Path,
    raw: Dict[str, Any],
    results: List[AllocationResult],
    summary: Dict[str, Decimal],
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row: int, cols: int):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # 摘要页
    ws_summary = wb.create_sheet("摘要")
    summary_rows = [
        ["指标", "值"],
        ["计算周期", raw.get("metadata", {}).get("period", "")],
        ["货币", raw.get("metadata", {}).get("currency", "USD")],
        ["GPU 总成本", float(summary["gpu_total_cost"])],
        ["Token 输入成本", float(summary["total_input_token_cost"])],
        ["Token 输出成本", float(summary["total_output_token_cost"])],
        ["Token 总成本", float(summary["total_token_cost"])],
        ["附加成本", float(summary["total_extra_cost"])],
        ["平台服务成本", float(summary["total_platform_cost"])],
        ["总成本", float(summary["total_cost"])],
        ["总 GPU 小时", float(summary["total_gpu_hours"])],
        ["总请求数", float(summary["total_requests"])],
        ["总 Token 数", float(summary["total_tokens"])],
    ]
    for row in summary_rows:
        ws_summary.append(row)
    style_header(ws_summary, 1, 2)
    for row_idx in range(2, ws_summary.max_row + 1):
        ws_summary.cell(row=row_idx, column=2).number_format = "#,##0.00"
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20

    # Workload 明细页
    ws_detail = wb.create_sheet("Workload 明细")
    detail_headers = [
        "ID", "Name", "Team", "Project", "Model", "Type",
        "GPU Hours", "Memory GB-Hours", "Requests",
        "Input Tokens", "Output Tokens", "Total Tokens",
        "GPU Cost", "Input Token Cost", "Output Token Cost",
        "Extra Cost", "Platform Cost", "Total Cost",
    ]
    ws_detail.append(detail_headers)
    style_header(ws_detail, 1, len(detail_headers))
    for r in results:
        ws_detail.append([
            r.workload.id,
            r.workload.name,
            r.workload.team,
            r.workload.project,
            r.workload.model_id,
            r.workload.workload_type,
            float(r.workload.gpu_hours),
            float(r.workload.memory_gb_hours),
            float(r.workload.requests),
            float(r.workload.input_tokens),
            float(r.workload.output_tokens),
            float(r.workload.total_tokens),
            float(r.gpu_allocated_cost),
            float(r.input_token_cost),
            float(r.output_token_cost),
            float(r.extra_cost),
            float(r.platform_allocated_cost),
            float(r.total_cost),
        ])
    for col in range(7, ws_detail.max_column + 1):
        col_letter = get_column_letter(col)
        for row_idx in range(2, ws_detail.max_row + 1):
            ws_detail.cell(row=row_idx, column=col).number_format = "#,##0.00"

    # 聚合页
    ws_alloc = wb.create_sheet("分摊汇总")
    ws_alloc.append(["维度", "键", "成本"])
    style_header(ws_alloc, 1, 3)
    for title, key_func in [
        ("Team", lambda r: r.workload.team),
        ("Project", lambda r: r.workload.project),
        ("Model", lambda r: r.workload.model_id),
    ]:
        groups: Dict[str, Decimal] = {}
        for r in results:
            key = key_func(r)
            groups[key] = groups.get(key, Decimal("0")) + r.total_cost
        for key, cost in sorted(groups.items(), key=lambda x: -x[1]):
            ws_alloc.append([title, key, float(cost)])
    for row_idx in range(2, ws_alloc.max_row + 1):
        ws_alloc.cell(row=row_idx, column=3).number_format = "#,##0.00"

    wb.save(str(xlsx_path))


def _write_excel_with_xlsxwriter(
    xlsx_path: Path,
    raw: Dict[str, Any],
    results: List[AllocationResult],
    summary: Dict[str, Decimal],
) -> None:
    workbook = xlsxwriter.Workbook(str(xlsx_path))
    bold = workbook.add_format({"bold": True})
    num_fmt = workbook.add_format({"num_format": "#,##0.00"})

    ws_summary = workbook.add_worksheet("摘要")
    summary_rows = [
        ["指标", "值"],
        ["计算周期", raw.get("metadata", {}).get("period", "")],
        ["货币", raw.get("metadata", {}).get("currency", "USD")],
        ["GPU 总成本", float(summary["gpu_total_cost"])],
        ["Token 输入成本", float(summary["total_input_token_cost"])],
        ["Token 输出成本", float(summary["total_output_token_cost"])],
        ["Token 总成本", float(summary["total_token_cost"])],
        ["附加成本", float(summary["total_extra_cost"])],
        ["平台服务成本", float(summary["total_platform_cost"])],
        ["总成本", float(summary["total_cost"])],
        ["总 GPU 小时", float(summary["total_gpu_hours"])],
        ["总请求数", float(summary["total_requests"])],
        ["总 Token 数", float(summary["total_tokens"])],
    ]
    for i, row in enumerate(summary_rows):
        ws_summary.write_row(i, 0, row, bold if i == 0 else None)
        if i > 0:
            ws_summary.write_number(i, 1, row[1], num_fmt) if isinstance(row[1], float) else None

    ws_detail = workbook.add_worksheet("Workload 明细")
    detail_headers = [
        "ID", "Name", "Team", "Project", "Model", "Type",
        "GPU Hours", "Memory GB-Hours", "Requests",
        "Input Tokens", "Output Tokens", "Total Tokens",
        "GPU Cost", "Input Token Cost", "Output Token Cost",
        "Extra Cost", "Platform Cost", "Total Cost",
    ]
    ws_detail.write_row(0, 0, detail_headers, bold)
    for i, r in enumerate(results, start=1):
        ws_detail.write_row(i, 0, [
            r.workload.id,
            r.workload.name,
            r.workload.team,
            r.workload.project,
            r.workload.model_id,
            r.workload.workload_type,
            float(r.workload.gpu_hours),
            float(r.workload.memory_gb_hours),
            float(r.workload.requests),
            float(r.workload.input_tokens),
            float(r.workload.output_tokens),
            float(r.workload.total_tokens),
            float(r.gpu_allocated_cost),
            float(r.input_token_cost),
            float(r.output_token_cost),
            float(r.extra_cost),
            float(r.platform_allocated_cost),
            float(r.total_cost),
        ], num_fmt)

    ws_alloc = workbook.add_worksheet("分摊汇总")
    ws_alloc.write_row(0, 0, ["维度", "键", "成本"], bold)
    row_idx = 1
    for title, key_func in [
        ("Team", lambda r: r.workload.team),
        ("Project", lambda r: r.workload.project),
        ("Model", lambda r: r.workload.model_id),
    ]:
        groups: Dict[str, Decimal] = {}
        for r in results:
            key = key_func(r)
            groups[key] = groups.get(key, Decimal("0")) + r.total_cost
        for key, cost in sorted(groups.items(), key=lambda x: -x[1]):
            ws_alloc.write_row(row_idx, 0, [title, key, float(cost)], num_fmt)
            row_idx += 1

    workbook.close()


def write_excel_report(
    output_base: Path,
    raw: Dict[str, Any],
    results: List[AllocationResult],
    summary: Dict[str, Decimal],
) -> Optional[Path]:
    xlsx_path = Path(str(output_base) + ".xlsx")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if HAS_OPENPYXL:
        _write_excel_with_openpyxl(xlsx_path, raw, results, summary)
    elif HAS_XLSXWRITER:
        _write_excel_with_xlsxwriter(xlsx_path, raw, results, summary)
    else:
        print(
            "警告：未检测到 openpyxl 或 xlsxwriter，已跳过 Excel 输出。"
            "请运行 'pip install openpyxl' 后重试。",
            file=sys.stderr,
        )
        return None

    return xlsx_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="AI GPU 成本分摊计算器",
    )
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="输入 YAML/JSON 文件路径",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("ai-gpu-cost-report"),
        help="输出文件基础路径（不含扩展名，默认：ai-gpu-cost-report）",
    )
    parser.add_argument(
        "--format",
        choices=["csv", "excel", "all"],
        default="all",
        help="输出格式（默认 all）",
    )
    args = parser.parse_args()

    raw = load_input(args.input)
    gpu_total_cost, results, summary = calculate(raw)

    print_report(raw, gpu_total_cost, results, summary)

    if args.format in {"csv", "all"}:
        csv_path = write_csv_report(args.output, results, summary)
        print(f"CSV 报告已写入: {csv_path}")

    if args.format in {"excel", "all"}:
        xlsx_path = write_excel_report(args.output, raw, results, summary)
        if xlsx_path:
            print(f"Excel 报告已写入: {xlsx_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
